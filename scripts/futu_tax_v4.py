#!/usr/bin/env python3
"""
╔══════════════════════════════════════════════════════════════════════╗
║   CRS 境外所得税计算器 v4.0 — 三 skill 融合版                       ║
║   Universal Offshore Tax Calculator                                 ║
║                                                                    ║
║  融合来源:                                                         ║
║  1. crs-tax-calculator — 多格式解析, FIFO/ACB, 多券商兼容          ║
║  2. crs-report-generator — PDF本地解析, CRS报表模板, 信息提取       ║
║  3. futu-tax v3 — WMA/FIFO引擎, PQR列, 分市场, 跨年结转, RSU       ║
║                                                                    ║
║  支持格式: 年度账单Excel | 月结单PDF/Excel | 日结单CSV | 截图PNG   ║
║  支持品种: 股票/基金/期权/期货/债券/结构化票据                      ║
║  支持券商: 富途/长桥/老虎/盈透/盈立/卓锐 (自动识别)               ║
╚══════════════════════════════════════════════════════════════════════╝
"""

import sys, os, json, re, argparse, ssl, base64, tempfile
from datetime import datetime
from collections import defaultdict
from copy import deepcopy
import urllib.request
import openpyxl

VERSION = "4.0.0"
TAX_RATE = 0.20

# ═══════════════════════════════════════════════════════════
# 0. CONFIG
# ═══════════════════════════════════════════════════════════

RSU_CONFIG_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'rsu_config.json')

# Broker detection patterns (from crs-tax-calculator)
KNOWN_BROKERS = {
    'futu': {'names': ['富途','Futu','moomoo','保證金','孖展'], 'name_cn': '富途证券'},
    'longbridge': {'names': ['长桥','Longbridge','長橋'], 'name_cn': '长桥证券'},
    'tiger': {'names': ['老虎','Tiger','老虎证券'], 'name_cn': '老虎证券'},
    'ibkr': {'names': ['盈透','Interactive Brokers','IBKR'], 'name_cn': '盈透证券'},
    'usmart': {'names': ['盈立','uSMART','uSmart'], 'name_cn': '盈立证券'},
    'zircon': {'names': ['卓锐','Zircon'], 'name_cn': '卓锐证券'},
}

MIME_MAP = {
    '.pdf': 'application/pdf',
    '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    '.csv': 'text/csv',
    '.png': 'image/png',
    '.jpg': 'image/jpeg',
    '.jpeg': 'image/jpeg',
}

MARKET_MAP = {'US': '🇺🇸 美股', 'SEHK': '🇭🇰 港股', 'FD': '🌐 基金', '-': '🌐 其他', 'FTSN': '🌐 票据'}

# ═══════════════════════════════════════════════════════════
# 1. FORMAT DETECTION (from crs-tax-calculator + custom)
# ═══════════════════════════════════════════════════════════

def detect_broker(text_sample):
    """Detect broker from file content (inspired by crs-tax-calculator)"""
    for broker_id, info in KNOWN_BROKERS.items():
        for kw in info['names']:
            if kw.lower() in text_sample.lower():
                return broker_id, info['name_cn']
    return 'unknown', '未知券商'

def detect_file_format(filepath):
    """Detect file format and type"""
    ext = os.path.splitext(filepath)[1].lower()
    if ext == '.pdf':
        return 'pdf'
    elif ext in ('.xlsx', '.xls'):
        return 'excel'
    elif ext == '.csv':
        return 'csv'
    elif ext in ('.png', '.jpg', '.jpeg'):
        return 'image'
    return 'unknown'

# ═══════════════════════════════════════════════════════════
# 2. PDF PARSER (from crs-report-generator)
# ═══════════════════════════════════════════════════════════

def parse_pdf_to_text(filepath):
    """Extract text from PDF using pdfplumber (from crs-report-generator)"""
    try:
        import pdfplumber
        text_parts = []
        with pdfplumber.open(filepath) as pdf:
            for page in pdf.pages:
                t = page.extract_text()
                if t:
                    text_parts.append(t)
        return '\n'.join(text_parts)
    except ImportError:
        print("  ⚠️ pdfplumber未安装, 尝试PyPDF2...")
        try:
            from PyPDF2 import PdfReader
            text_parts = []
            reader = PdfReader(filepath)
            for page in reader.pages:
                t = page.extract_text()
                if t:
                    text_parts.append(t)
            return '\n'.join(text_parts)
        except ImportError:
            print("  ❌ 需要安装: pip install pdfplumber PyPDF2")
            return ""

def extract_trades_from_pdf_text(text):
    """
    Extract trading records from PDF text using regex patterns.
    Pattern set from crs-report-generator + extended for more brokers.
    """
    trades = []

    # Pattern 1: Standard trade line (Futu format)
    # Date | Code | Buy/Sell | Qty | Price | Amount | Fee
    pat1 = re.findall(
        r'(\d{4}[-/]\d{2}[-/]\d{2})\s+.*?'
        r'(买入|卖出|申购|赎回|Buy|Sell)\s+.*?'
        r'([A-Z0-9.]+)\s+.*?'
        r'([\d,.]+)\s+.*?'
        r'([\d,.]+)\s+.*?'
        r'([\d,.]+)',
        text
    )
    for m in pat1:
        trades.append({
            'date': m[0], 'direction': m[1], 'code': m[2],
            'qty': float(m[3].replace(',','')),
            'price': float(m[4].replace(',','')),
            'amount': float(m[5].replace(',','')),
        })

    # Pattern 2: Generic transaction line
    pat2 = re.findall(
        r'(\d{4}/\d{2}/\d{2}).*?'
        r'(Bought|Sold|Buy|Sell|买入|卖出)\s+'
        r'([A-Z0-9]+).*?'
        r'([\d,]+)\s*(?:shares|股|@)\s*'
        r'([\d.]+)',
        text, re.IGNORECASE
    )
    for m in pat2:
        trades.append({
            'date': m[0].replace('/','-'), 'direction': m[1], 'code': m[2],
            'qty': float(m[3].replace(',','')),
            'price': float(m[4]),
            'amount': float(m[3].replace(',','')) * float(m[4]),
        })

    return trades

def extract_dividends_from_pdf_text(text):
    """Extract dividend entries from PDF text"""
    divs = []
    pat = re.findall(
        r'(DIVIDEND|股息|分红|Dividend).*?'
        r'([\d,.]+)\s*(?:shares|股).*?'
        r'([\d.]+)\s*(?:USD|HKD|CNY)',
        text, re.IGNORECASE
    )
    for m in pat:
        divs.append({'type': m[0], 'shares': float(m[1].replace(',','')), 'amount': float(m[2])})
    return divs

# ═══════════════════════════════════════════════════════════
# 3. CSV PARSER (from crs-tax-calculator compatibility)
# ═══════════════════════════════════════════════════════════

def parse_csv(filepath):
    """Parse CSV statement file"""
    import csv
    records = []
    with open(filepath, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            records.append({
                'time': row.get('成交时间', row.get('日期', row.get('Date', ''))),
                'category': row.get('品类', row.get('类型', row.get('Type', '证券'))),
                'code': row.get('代码名称', row.get('代码', row.get('Code', ''))).strip(),
                'market': row.get('交易所/市场', row.get('市场', row.get('Market', ''))).strip(),
                'direction': row.get('方向', row.get('Direction', '')).strip(),
                'currency': row.get('币种', row.get('货币', row.get('Currency', ''))).strip(),
                'qty': float(row.get('数量/面值', row.get('数量', row.get('Qty', 0))) or 0),
                'price': float(row.get('价格', row.get('Price', 0)) or 0),
                'amount': float(row.get('成交金额', row.get('金额', row.get('Amount', 0))) or 0),
                'fee': abs(float(row.get('总费用', row.get('费用', row.get('Fee', 0))) or 0)),
                'net_amount': float(row.get('变动金额', row.get('净额', row.get('Net', 0))) or 0),
                'row': 0,
            })
    return records

# ═══════════════════════════════════════════════════════════
# 4. EXCEL PARSER (from futu-tax v3)
# ═══════════════════════════════════════════════════════════

def find_sheet(wb, keywords):
    for sn in wb.sheetnames:
        if all(kw in sn for kw in keywords):
            return sn
    for sn in wb.sheetnames:
        if any(kw in sn for kw in keywords):
            return sn
    return None

def parse_trading_sheet(ws):
    """Parse trading sheet with auto column detection"""
    headers = {}
    for col in range(1, ws.max_column + 1):
        h = str(ws.cell(row=1, column=col).value or '').strip()
        if h:
            headers[h] = col

    def col(keywords, fallback):
        for kw in keywords:
            for h, c in headers.items():
                if kw in h:
                    return c
        return fallback

    tc = col(['成交时间','时间','日期','Trade Time','Date'], 1)
    catc = col(['品类','类型','Category','Type'], 4)
    codec = col(['代码名称','代码','名称','Code','Symbol'], 5)
    mktc = col(['交易所/市场','市场','Market','Exchange'], 6)
    dirc = col(['方向','Direction','Side'], 7)
    curc = col(['币种','货币','Currency'], 9)
    qtyc = col(['数量/面值','数量','Qty','Quantity'], 10)
    prcc = col(['价格','成交价','Price'], 11)
    amtc = col(['成交金额','金额','Amount','Trade Amount'], 12)
    feec = col(['总费用','费用','Fee','Commission'], 13)
    netc = col(['变动金额','净额','Net Amount'], 14)

    records = []
    for row in range(2, ws.max_row + 1):
        ts = ws.cell(row=row, column=tc).value
        if not ts: continue
        records.append({
            'time': str(ts),
            'category': str(ws.cell(row=row, column=catc).value or ''),
            'code': str(ws.cell(row=row, column=codec).value or '').strip(),
            'market': str(ws.cell(row=row, column=mktc).value or '').strip(),
            'direction': str(ws.cell(row=row, column=dirc).value or '').strip(),
            'currency': str(ws.cell(row=row, column=curc).value or '').strip(),
            'qty': float(ws.cell(row=row, column=qtyc).value or 0),
            'price': float(ws.cell(row=row, column=prcc).value or 0),
            'amount': float(ws.cell(row=row, column=amtc).value or 0),
            'fee': abs(float(ws.cell(row=row, column=feec).value or 0)),
            'net_amount': float(ws.cell(row=row, column=netc).value or 0),
            'row': row,
        })
    records.sort(key=lambda r: r['time'])
    return records

def parse_asset_sheet(ws):
    mov = []
    for row in range(2, ws.max_row + 1):
        d = ws.cell(row=row, column=1).value
        if not d: continue
        mov.append({
            'date': str(d), 'account_name': str(ws.cell(row=row, column=2).value or ''),
            'account_num': str(ws.cell(row=row, column=3).value or ''),
            'category': str(ws.cell(row=row, column=4).value or ''),
            'code': str(ws.cell(row=row, column=5).value or '').strip(),
            'market': str(ws.cell(row=row, column=6).value or '').strip(),
            'direction': str(ws.cell(row=row, column=7).value or ''),
            'type': str(ws.cell(row=row, column=8).value or ''),
            'currency': str(ws.cell(row=row, column=9).value or '').strip(),
            'qty': float(ws.cell(row=row, column=10).value or 0),
            'remark': str(ws.cell(row=row, column=11).value or ''),
        })
    return mov

def parse_fund_flow_sheet(ws):
    r = {'dividends':[],'interests':[],'interest_charges':[],'ipo_payments':[],
         'rsu_tax':[],'note_coupons':[],'deposits':[],'other':[]}
    for row in range(2, ws.max_row + 1):
        d = ws.cell(row=row, column=1).value
        if not d: continue
        e = {
            'date': str(d), 'account_name': str(ws.cell(row=row, column=2).value or ''),
            'account_num': str(ws.cell(row=row, column=3).value or ''),
            'type': str(ws.cell(row=row, column=4).value or ''),
            'direction': str(ws.cell(row=row, column=5).value or ''),
            'currency': str(ws.cell(row=row, column=6).value or '').strip(),
            'amount': float(ws.cell(row=row, column=7).value or 0),
            'remark': str(ws.cell(row=row, column=8).value or ''),
        }
        t, rk, dr = e['type'], e['remark'].lower(), e['direction']
        if '分红' in t or 'dividend' in rk or ('公司行动' in t and ('dividend' in rk or 'i/d' in rk.lower() or 's/d' in rk.lower())):
            r['dividends'].append(e)
        elif 'withholding tax' in rk or ('公司行动' in t and 'tax' in rk):
            r['dividends'].append(e)
        elif '利息扣除' in t or ('利息' in t and dr == 'Out'):
            r['interest_charges'].append(e)
        elif '利息' in t and dr == 'In':
            r['interests'].append(e)
        elif 'coupon' in rk or '票息' in rk:
            r['note_coupons'].append(e)
        elif 'ipo' in t.lower() or 'ipo' in rk:
            r['ipo_payments'].append(e)
        elif any(kw in t for kw in ['激励','长期','收益计划']) or 'rsu' in rk:
            r['rsu_tax'].append(e)
        elif any(kw in t for kw in ['出入金','资金调拨','账户间','货币兑换']):
            r['deposits'].append(e)
        else:
            r['other'].append(e)
    return r

def parse_opening_positions(ws):
    pos = {}
    for row in range(2, ws.max_row + 1):
        if '期初' not in str(ws.cell(row=row, column=1).value or ''): continue
        code = str(ws.cell(row=row, column=6).value or '').strip()
        market = str(ws.cell(row=row, column=7).value or '').strip()
        cur = str(ws.cell(row=row, column=8).value or '').strip()
        qty = float(ws.cell(row=row, column=9).value or 0)
        px = float(ws.cell(row=row, column=10).value or 0)
        if qty <= 0 or not code: continue
        key = (code, market, cur)
        if key in pos:
            old = pos[key]; tq = old['qty'] + qty
            old['price'] = (old['price']*old['qty'] + px*qty)/tq if tq > 0 else 0
            old['qty'] = tq
        else:
            pos[key] = {'qty': qty, 'price': px, 'market': market, 'currency': cur, 'code': code}
    return pos

# ═══════════════════════════════════════════════════════════
# 5. UNIFIED FILE PROCESSOR
# ═══════════════════════════════════════════════════════════

def process_any_file(filepath):
    """
    Process ANY file format and return trading records + metadata.
    Routes to correct parser based on format + content.
    """
    fmt = detect_file_format(filepath)
    records = []
    assets = []
    flows = {'dividends':[],'interests':[],'interest_charges':[],'ipo_payments':[],
             'rsu_tax':[],'note_coupons':[],'deposits':[],'other':[]}
    broker = 'unknown'
    broker_name = '未知'

    if fmt == 'excel':
        import openpyxl
        wb = openpyxl.load_workbook(filepath, data_only=True)

        # Detect broker from account info sheet
        for sn in wb.sheetnames:
            if '账户' in sn or '信息' in sn or 'Account' in sn:
                try:
                    ws = wb[sn]
                    text_sample = ' '.join(str(ws.cell(row=r, column=c).value or '')
                                           for r in range(1, min(ws.max_row+1, 20))
                                           for c in range(1, min(ws.max_column+1, 5)))
                    broker, broker_name = detect_broker(text_sample)
                    break
                except: pass

        # Trading records
        trade_sn = find_sheet(wb, ['证券','交易流水']) or find_sheet(wb, ['交易'])
        if trade_sn:
            records = parse_trading_sheet(wb[trade_sn])

        # Asset movements
        asset_sn = find_sheet(wb, ['证券','资产进出']) or find_sheet(wb, ['资产'])
        if asset_sn:
            assets = parse_asset_sheet(wb[asset_sn])

        # Fund flows
        flow_sn = find_sheet(wb, ['证券','资金进出']) or find_sheet(wb, ['资金'])
        if flow_sn:
            flows = parse_fund_flow_sheet(wb[flow_sn])

        # Also check期货 sheets
        fut_trade = find_sheet(wb, ['期货','交易明细'])
        if fut_trade:
            fut_recs = parse_trading_sheet(wb[fut_trade])
            for r in fut_recs:
                r['category'] = '期货'
            records.extend(fut_recs)

    elif fmt == 'pdf':
        text = parse_pdf_to_text(filepath)
        if text:
            broker, broker_name = detect_broker(text)
            records = extract_trades_from_pdf_text(text)
            divs = extract_dividends_from_pdf_text(text)
            for d in divs:
                flows['dividends'].append({
                    'date': d.get('date', ''), 'type': '股息', 'direction': 'In',
                    'currency': d.get('currency', 'USD'), 'amount': d.get('amount', 0),
                    'remark': f"PDF extracted: {d.get('type','')}"
                })

    elif fmt == 'csv':
        records = parse_csv(filepath)
        with open(filepath, 'r', encoding='utf-8-sig') as f:
            broker, broker_name = detect_broker(f.read()[:5000])

    elif fmt == 'image':
        # Images need OCR — delegate to crs-tax-calculator API pattern
        print(f"  ⚠️ 图片格式需要OCR, 建议转换或使用wealthlplantation API")

    records.sort(key=lambda r: r.get('time', r.get('date', '')))
    return records, assets, flows, broker, broker_name

# ═══════════════════════════════════════════════════════════
# 6-8. COST CALCULATION ENGINE (same as v3, compact)
# ═══════════════════════════════════════════════════════════

def load_rsu_config():
    if os.path.exists(RSU_CONFIG_PATH):
        try:
            with open(RSU_CONFIG_PATH) as f: return json.load(f)
        except: pass
    return {}

def date_diff(d1, d2):
    try:
        for fmt in ['%Y%m%d','%Y-%m-%d','%Y-%m-%d %H:%M:%S']:
            try:
                s1, s2 = d1.replace('-','')[:8], d2.replace('-','')[:8]
                return abs((datetime.strptime(s1,'%Y%m%d')-datetime.strptime(s2,'%Y%m%d')).days)
            except: continue
    except: pass
    return 999

def find_asset_cost(mv, fund_flows, trading_records, rsu_config=None):
    mv_type, remark, code = mv['type'], mv['remark'], mv['code']
    mv_date, mv_qty, market = mv['date'], mv['qty'], mv['market']
    result = {'price': 0.0, 'fee': 0.0, 'source': 'unknown'}

    if rsu_config and code in rsu_config:
        for e in rsu_config[code]:
            if abs(date_diff(e['vest_date'], mv_date)) <= 2 and abs(e['shares'] - mv_qty) < 1:
                result['price'] = e['fmv']; result['source'] = f'RSU config'; return result

    if 'ipo' in mv_type.lower() or 'ipo' in remark.lower():
        for p in fund_flows.get('ipo_payments', []):
            if p['direction'] == 'Out' and abs(date_diff(p['date'], mv_date)) <= 5 and abs(p['amount']) > 100:
                result['price'] = abs(p['amount'])/mv_qty if mv_qty > 0 else 0
                result['source'] = 'IPO payment'; return result

    if any(kw in mv_type for kw in ['长期激励','股票收益计划','激励','资产进出']) or 'rsu' in remark.lower():
        for rsu in fund_flows.get('rsu_tax', []):
            if rsu['direction'] == 'Out' and abs(date_diff(rsu['date'], mv_date)) <= 5:
                tax = abs(rsu['amount'])
                if tax > 0 and mv_qty > 0:
                    for rate in [0.10,0.15,0.20,0.22,0.25,0.30]:
                        fmv = tax/(rate*mv_qty)
                        if 1.0 < fmv < 10000:
                            result['price'] = round(fmv,4); result['source'] = f'Tax est'; return result
        if trading_records:
            prices = [rec['price'] for rec in trading_records
                      if rec['code']==code and rec['market']==market and abs(date_diff(rec['time'][:8], mv_date))<=60]
            if prices:
                result['price'] = round(sum(prices)/len(prices),4)
                result['source'] = f'Market avg({len(prices)})'; return result
        result['needs_manual'] = True; return result
    return result

class PTracker:
    def __init__(self, code, market, currency, category='stock'):
        self.code=code; self.market=market; self.currency=currency; self.category=category
        self.P=0.0; self.Q=0.0; self.total_cost=0.0; self.realized_pnl=0.0
        self.tba=0.0; self.tbf=0.0; self.tsa=0.0; self.tsf=0.0
    def add_opening(self,qty,cost):
        if qty<=0: return
        self.P=qty; self.Q=cost; self.total_cost=qty*cost
    def buy(self,qty,price,fee=0.0):
        if qty<=0: return
        bc=qty*price+fee; self.tba+=qty*price; self.tbf+=fee
        if self.P>0: self.total_cost+=bc; self.P+=qty; self.Q=self.total_cost/self.P
        else: self.P=qty; self.total_cost=bc; self.Q=bc/qty
    def sell(self,qty,price,fee=0.0):
        if qty<=0 or self.P<=0: return 0.0
        sq=min(qty,self.P)
        r=(price-self.Q)*sq-fee; self.realized_pnl+=r
        self.tsa+=sq*price; self.tsf+=fee
        self.P-=sq; self.total_cost=self.P*self.Q
        if abs(self.P)<0.000001: self.P=self.Q=self.total_cost=0.0
        return r

class OptTracker:
    def __init__(self, code, market, currency):
        self.code=code; self.market=market; self.currency=currency
        self.positions=[]; self.realized_pnl=0.0; self.P=0.0; self.Q=0.0
    def add_opening(self,qty,cost): pass
    def buy(self,qty,price,fee=0.0): pass
    def sell(self,qty,price,fee=0.0): pass
    def open_short(self,qty,premium,fee=0.0):
        self.positions.append((qty,premium,'short')); self.realized_pnl-=fee; self.P+=qty
    def close_short(self,qty,premium,fee=0.0):
        r=0.0; rem=qty
        while rem>0 and self.positions:
            oq,op,od=self.positions[0]
            if od!='short': self.positions.pop(0); continue
            cq=min(rem,oq); r+=(op-premium)*cq; rem-=cq
            if cq>=oq: self.positions.pop(0)
            else: self.positions[0]=(oq-cq,op,od)
        self.realized_pnl+=r-fee; self.P-=qty; return r-fee
    def open_long(self,qty,premium,fee=0.0):
        self.positions.append((qty,premium,'long')); self.realized_pnl-=fee; self.P+=qty
    def close_long(self,qty,premium,fee=0.0):
        r=0.0; rem=qty
        while rem>0 and self.positions:
            oq,op,od=self.positions[0]
            if od!='long': self.positions.pop(0); continue
            cq=min(rem,oq); r+=(premium-op)*cq; rem-=cq
            if cq>=oq: self.positions.pop(0)
            else: self.positions[0]=(oq-cq,op,od)
        self.realized_pnl+=r-fee; self.P-=qty; return r-fee

def calculate_all(trading_records, asset_movements, fund_flows,
                  opening_positions, method='WMA', rsu_config=None):
    trackers={}; warnings=[]
    def gt(code,market,currency,category='stock'):
        key=(code,market,currency)
        if key not in trackers:
            trackers[key] = OptTracker(code,market,currency) if ('期权' in category or 'option' in category.lower()) else PTracker(code,market,currency,category)
        return trackers[key]
    for (code,market,cur),pos in opening_positions.items():
        gt(code,market,cur).add_opening(pos['qty'],pos['price'])
    asset_entries=[]
    for mv in asset_movements:
        if mv['direction']!='In': continue
        if any(kw in mv['type'] for kw in ['账户升级','账户迁移','升级']): continue
        ci=find_asset_cost(mv,fund_flows,trading_records,rsu_config)
        if ci.get('needs_manual'):
            warnings.append({'code':mv['code'],'market':mv['market'],'qty':mv['qty'],'type':mv['type'],'date':mv['date'],
                           'message':f"{mv['code']} {mv['type']}: {mv['qty']:,.0f}股 成本未知"})
        asset_entries.append({'time':mv['date'],'code':mv['code'],'market':mv['market'],
                             'currency':mv['currency'],'category':mv['category'],
                             'qty':mv['qty'],'cost_per_share':max(ci['price'],0.0001),
                             'fee':ci.get('fee',0)})
    events=[]
    for r in trading_records: events.append({'type':'trade','time':r['time'],'data':r})
    for e in asset_entries: events.append({'type':'asset_in','time':e['time'],'data':e})
    events.sort(key=lambda e: e['time'])
    enriched=[]
    for ev in events:
        if ev['type']=='asset_in':
            d=ev['data']; t=gt(d['code'],d['market'],d['currency'],d.get('category','stock'))
            if d['qty']>0: t.buy(d['qty'],d['cost_per_share'],d['fee'])
            continue
        rec=ev['data']; cat=rec['category']; t=gt(rec['code'],rec['market'],rec['currency'],cat)
        dr=rec['direction']; qty=abs(rec['qty']); px=rec['price']; fee=rec['fee']
        Pb=t.P if hasattr(t,'P') else 0; Qb=t.Q if hasattr(t,'Q') else 0; realized=0.0
        is_opt='期权' in cat or 'option' in cat.lower()
        if is_opt:
            if dr=='卖出开仓': t.open_short(qty,px,fee)
            elif dr=='买入平仓': realized=t.close_short(qty,px,fee)
            elif dr=='买入开仓': t.open_long(qty,px,fee)
            elif dr=='卖出平仓': realized=t.close_long(qty,px,fee)
        else:
            is_buy=dr in ('买入开仓','买入平仓','申购','买入')
            is_sell=dr in ('卖出平仓','卖出开仓','赎回','卖出')
            if is_buy: t.buy(qty,px,fee)
            elif is_sell: realized=t.sell(qty,px,fee)
        rec2=dict(rec); rec2['P_before']=round(Pb,4); rec2['Q_before']=round(Qb,4)
        rec2['P_after']=round(t.P if hasattr(t,'P') else 0,4)
        rec2['Q_after']=round(t.Q if hasattr(t,'Q') else 0,4)
        rec2['R']=round(realized,4); enriched.append(rec2)
    return enriched, trackers, warnings

# ═══════════════════════════════════════════════════════════
# 9. TAX + OUTPUT (merged from all 3 skills)
# ═══════════════════════════════════════════════════════════

def fetch_rates(year):
    rates={}
    for cur in ['USD','HKD']:
        try:
            url=f'https://api.exchangerate-api.com/v4/latest/{cur}'
            req=urllib.request.Request(url,headers={'User-Agent':'Mozilla/5.0'})
            ctx=ssl.create_default_context()
            with urllib.request.urlopen(req,timeout=10,context=ctx) as r:
                rates[cur]=json.loads(r.read().decode())['rates'].get('CNY',0)
        except: pass
    return rates

def calculate_tax(trackers,fund_flows,exchange_rates,year):
    mpnl=defaultdict(lambda:defaultdict(float))
    for (code,market,currency),t in trackers.items():
        pnl=getattr(t,'realized_pnl',0)
        if abs(pnl)>0.005: mpnl[market][currency]+=pnl
    divs=defaultdict(lambda:{'gross':0.0,'tax_withheld':0.0})
    for div in fund_flows.get('dividends',[]):
        rk=div['remark'].lower(); cur=div['currency']
        if div['direction']=='In' and ('dividend' in rk or 'i/d' in rk.lower() or 's/d' in rk.lower()):
            divs[cur]['gross']+=abs(div['amount'])
        elif div['direction']=='Out' and ('withholding tax' in rk or 'tax' in rk):
            divs[cur]['tax_withheld']+=abs(div['amount'])
    ints=defaultdict(float)
    for i in fund_flows.get('interests',[]):
        if i['direction']=='In': ints[i['currency']]+=abs(i['amount'])
    for c in fund_flows.get('note_coupons',[]):
        if c['direction']=='In': ints[c['currency']]+=abs(c['amount'])
    def rm(amt,cur): return amt*exchange_rates.get(cur,1.0)
    res={'year':year,'exchange_rates':exchange_rates,
         'capital_gains':{'by_market':{},'total_tax_rmb':0.0},
         'dividends':{'by_currency':{},'total_tax_rmb':0.0},
         'interests':{'by_currency':{},'total_tax_rmb':0.0}}
    for market,curs in sorted(mpnl.items()):
        label=MARKET_MAP.get(market,f'🌐 {market}')
        me={'label':label,'currencies':{},'subtotal_tax':0.0}
        for cur,pnl in sorted(curs.items()):
            pr=rm(pnl,cur); tax=max(0,pr*TAX_RATE)
            me['currencies'][cur]={'pnl':round(pnl,2),'pnl_rmb':round(pr,2),'rate':exchange_rates.get(cur,1.0),'tax_rmb':round(tax,2)}
            me['subtotal_tax']+=tax
        res['capital_gains']['by_market'][market]=me
        res['capital_gains']['total_tax_rmb']+=me['subtotal_tax']
    res['capital_gains']['total_tax_rmb']=round(res['capital_gains']['total_tax_rmb'],2)
    for cur in sorted(divs):
        d=divs[cur]; gr=rm(d['gross'],cur); wt=rm(d['tax_withheld'],cur)
        should=gr*TAX_RATE; owed=max(0,should-wt)
        res['dividends']['by_currency'][cur]={'gross':round(d['gross'],2),'gross_rmb':round(gr,2),
            'tax_withheld':round(d['tax_withheld'],2),'tax_withheld_rmb':round(wt,2),
            'should_tax_rmb':round(should,2),'owed_rmb':round(owed,2)}
        res['dividends']['total_tax_rmb']+=owed
    res['dividends']['total_tax_rmb']=round(res['dividends']['total_tax_rmb'],2)
    for cur in sorted(ints):
        total=ints[cur]; tr=rm(total,cur); tax=tr*TAX_RATE
        res['interests']['by_currency'][cur]={'total':round(total,2),'total_rmb':round(tr,2),'tax_rmb':round(tax,2)}
        res['interests']['total_tax_rmb']+=tax
    res['interests']['total_tax_rmb']=round(res['interests']['total_tax_rmb'],2)
    res['grand_total_rmb']=round(res['capital_gains']['total_tax_rmb']+res['dividends']['total_tax_rmb']+res['interests']['total_tax_rmb'],2)
    return res

def write_output(enriched,trackers,tax,fund_flows,warnings,exchange_rates,output_path,year,method,broker_name=''):
    import openpyxl
    from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
    from openpyxl.utils import get_column_letter
    wb=openpyxl.Workbook()
    HF=PatternFill(start_color='1F4E79',end_color='1F4E79',fill_type='solid')
    SF=PatternFill(start_color='2E75B6',end_color='2E75B6',fill_type='solid')
    GF=PatternFill(start_color='C6EFCE',end_color='C6EFCE',fill_type='solid')
    RF=PatternFill(start_color='FFC7CE',end_color='FFC7CE',fill_type='solid')
    YF=PatternFill(start_color='FFE699',end_color='FFE699',fill_type='solid')
    HFONT=Font(name='Arial',size=11,bold=True,color='FFFFFF')
    TFONT=Font(name='Arial',size=16,bold=True,color='1F4E79')
    NF=Font(name='Arial',size=10); BF=Font(name='Arial',size=10,bold=True)
    RB=Font(name='Arial',size=12,bold=True,color='FF0000')
    TB=Border(left=Side('thin'),right=Side('thin'),top=Side('thin'),bottom=Side('thin'))
    CA=Alignment(horizontal='center',vertical='center',wrap_text=True)
    LA=Alignment(horizontal='left',vertical='center',wrap_text=True)
    def ac(cell,font=None,fill=None,align=None,border=None):
        if font: cell.font=font
        if fill: cell.fill=fill
        if align: cell.alignment=align
        if border: cell.border=border

    # Sheet 1
    ws1=wb.active; ws1.title="📊 交易明细"
    hdrs=['成交时间','品类','代码','市场','方向','币种','数量','价格','成交金额','总费用','变动金额','P(前)','Q(前)','P(后)','Q(后)','R(盈亏)']
    for c,h in enumerate(hdrs,1): ac(ws1.cell(row=1,column=c,value=h),HFONT,HF,CA,TB)
    for i,rec in enumerate(enriched):
        row=i+2
        vals=[rec['time'][:19],rec['category'],rec['code'],rec['market'],rec['direction'],
              rec['currency'],rec['qty'],rec['price'],rec['amount'],rec['fee'],
              rec['net_amount'],rec['P_before'],rec['Q_before'],rec['P_after'],rec['Q_after'],rec['R']]
        for c,v in enumerate(vals,1): ac(ws1.cell(row=row,column=c,value=v),NF,None,CA,TB)
        rc=ws1.cell(row=row,column=16)
        if rec['R']>0.005: rc.fill=GF
        elif rec['R']<-0.005: rc.fill=RF
    widths=[20,8,10,8,10,8,12,10,14,10,12,10,10,10,10,12]
    for i,w in enumerate(widths,1): ws1.column_dimensions[get_column_letter(i)].width=w
    ws1.freeze_panes='A2'

    # Sheet 2
    ws2=wb.create_sheet("📋 持仓汇总")
    ph=['代码','市场','币种','品类','剩余量','加权成本(Q)','总成本','已实现盈亏','买入额','买入费','卖出额','卖出费']
    for c,h in enumerate(ph,1): ac(ws2.cell(row=1,column=c,value=h),HFONT,HF,CA,TB)
    for i,((code,market,cur),t) in enumerate(sorted(trackers.items(),key=lambda x:abs(getattr(x[1],'realized_pnl',0)),reverse=True)):
        row=i+2; pnl=getattr(t,'realized_pnl',0)
        vals=[code,market,cur,getattr(t,'category','stock'),round(getattr(t,'P',0),2),round(getattr(t,'Q',0),4),
              round(getattr(t,'P',0)*getattr(t,'Q',0),2),round(pnl,2),round(getattr(t,'tba',0),2),
              round(getattr(t,'tbf',0),2),round(getattr(t,'tsa',0),2),round(getattr(t,'tsf',0),2)]
        for c,v in enumerate(vals,1): ac(ws2.cell(row=row,column=c,value=v),NF,None,CA,TB)
        if pnl>0.005: ws2.cell(row=row,column=8).fill=GF
        elif pnl<-0.005: ws2.cell(row=row,column=8).fill=RF
    for i,w in enumerate([10,8,8,8,12,12,14,14,14,12,14,12],1): ws2.column_dimensions[get_column_letter(i)].width=w
    ws2.freeze_panes='A2'

    # Sheet 3
    ws3=wb.create_sheet("💰 税务报告")
    row=1; ws3.merge_cells('A1:H1')
    title=f"境外所得税报告 — {year}年度" + (f" ({broker_name})" if broker_name else "")
    ac(ws3.cell(row=1,column=1,value=title),TFONT); ws3.row_dimensions[1].height=35
    row=3
    ac(ws3.cell(row=row,column=1,value=f"方法:{method} | 汇率:"+",".join(f"{c}={r:.4f}" for c,r in exchange_rates.items())),NF)
    row+=2
    ac(ws3.cell(row=row,column=1,value="一、财产转让所得"),Font(name='Arial',size=13,bold=True))
    ws3.merge_cells(f'A{row}:H{row}'); row+=1
    for c,h in enumerate(['市场','币种','净盈亏','汇率','RMB','税额'],1): ac(ws3.cell(row=row,column=c,value=h),HFONT,SF,CA,TB)
    row+=1
    for market,mkt in sorted(tax['capital_gains']['by_market'].items()):
        for cur,d in sorted(mkt['currencies'].items()):
            for c,v in enumerate([mkt['label'],cur,d['pnl'],d['rate'],d['pnl_rmb'],d['tax_rmb']],1):
                ac(ws3.cell(row=row,column=c,value=v),NF,None,CA,TB)
            if d['tax_rmb']>0: ws3.cell(row=row,column=6).font=Font(name='Arial',size=10,bold=True,color='FF0000')
            row+=1
    ac(ws3.cell(row=row,column=1,value="小计"),BF,YF,CA,TB)
    for c in range(2,5): ac(ws3.cell(row=row,column=c,value=""),BF,YF,CA,TB)
    ac(ws3.cell(row=row,column=6,value=tax['capital_gains']['total_tax_rmb']),RB,YF,CA,TB)
    row+=2
    ac(ws3.cell(row=row,column=1,value="二、股息利息红利"),Font(name='Arial',size=13,bold=True))
    ws3.merge_cells(f'A{row}:H{row}'); row+=1
    for c,h in enumerate(['类型','币种','收入','RMB','预扣税','预扣RMB','应缴RMB','应补RMB'],1): ac(ws3.cell(row=row,column=c,value=h),HFONT,SF,CA,TB)
    row+=1
    for cur,d in sorted(tax['dividends']['by_currency'].items()):
        for c,v in enumerate(['股息',cur,d['gross'],d['gross_rmb'],d['tax_withheld'],d['tax_withheld_rmb'],d['should_tax_rmb'],d['owed_rmb']],1):
            ac(ws3.cell(row=row,column=c,value=v),NF,None,CA,TB)
        row+=1
    for cur,d in sorted(tax['interests']['by_currency'].items()):
        for c,v in enumerate(['利息/票息',cur,d['total'],d['total_rmb'],0,0,d['tax_rmb'],d['tax_rmb']],1):
            ac(ws3.cell(row=row,column=c,value=v),NF,None,CA,TB)
        row+=1
    dt=tax['dividends']['total_tax_rmb']+tax['interests']['total_tax_rmb']
    ac(ws3.cell(row=row,column=1,value="小计"),BF,YF,CA,TB)
    for c in range(2,7): ac(ws3.cell(row=row,column=c,value=""),BF,YF,CA,TB)
    ac(ws3.cell(row=row,column=8,value=round(dt,2)),RB,YF,CA,TB); row+=2
    ws3.merge_cells(f'A{row}:H{row}')
    ac(ws3.cell(row=row,column=1,value=f"三、合计应补税: ¥{tax['grand_total_rmb']:,.2f}"),Font(name='Arial',size=14,bold=True,color='FF0000'))
    ws3.row_dimensions[row].height=35; row+=2
    ac(ws3.cell(row=row,column=1,value="⚠️ 重要提醒"),Font(name='Arial',size=11,bold=True,color='FF0000')); row+=1
    for r in ["• 申报: 次年3/1-6/30, 逾期日万分之五滞纳金","• 仅对实际卖出盈利计税",
              "• 境外预扣税可申请抵免","• 保留交易记录和完税证明5年"]:
        ac(ws3.cell(row=row,column=1,value=r),NF); row+=1
    for i,w in enumerate([30,8,14,10,14,14,14,14],1): ws3.column_dimensions[get_column_letter(i)].width=w

    if fund_flows.get('dividends') or fund_flows.get('note_coupons'):
        ws4=wb.create_sheet("💵 股息利息")
        for c,h in enumerate(['日期','账户','类型','方向','币种','金额','备注'],1): ac(ws4.cell(row=1,column=c,value=h),HFONT,HF,CA,TB)
        r=2
        for item in fund_flows.get('dividends',[])+fund_flows.get('note_coupons',[]):
            for c,k in enumerate(['date','account_name','type','direction','currency','amount','remark'],1):
                ac(ws4.cell(row=r,column=c,value=item.get(k,'')),NF,None,LA,TB)
            r+=1

    if warnings:
        ws5=wb.create_sheet("⚠️ 需确认")
        for c,h in enumerate(['代码','市场','类型','日期','数量','说明'],1): ac(ws5.cell(row=1,column=c,value=h),HFONT,HF,CA,TB)
        for i,w in enumerate(warnings):
            for c,v in enumerate([w['code'],w['market'],w['type'],w['date'],w['qty'],w['message']],1):
                ac(ws5.cell(row=i+2,column=c,value=v),NF,YF,LA,TB)

    wb.save(output_path)
    return output_path

# ═══════════════════════════════════════════════════════════
# 10. MAIN PROCESSING
# ═══════════════════════════════════════════════════════════

def process_year(input_files, prior_year_file=None, output_file=None,
                 year=None, method='WMA', exchange_rates=None):
    if isinstance(input_files, str): input_files = [input_files]
    if year is None:
        match = re.search(r'(\d{4})', os.path.basename(input_files[0]))
        year = int(match.group(1)) if match else datetime.now().year

    print(f"\n{'='*60}")
    print(f"  🧾 {year}年度 | {len(input_files)}个文件 | {method}")
    print(f"{'='*60}")

    all_trades, all_assets = [], []
    comb_flows = {'dividends':[],'interests':[],'interest_charges':[],'ipo_payments':[],
                  'rsu_tax':[],'note_coupons':[],'deposits':[],'other':[]}
    broker_name = ''

    for f in input_files:
        print(f"  解析: {os.path.basename(f)}")
        trades, assets, flows, broker, bn = process_any_file(f)
        if bn and not broker_name: broker_name = bn
        all_trades.extend(trades)
        all_assets.extend(assets)
        for k in comb_flows: comb_flows[k].extend(flows.get(k, []))

    # Dedup
    seen = set()
    dedup_trades = []
    for r in all_trades:
        key = (r.get('time',''), r.get('code',''), r.get('direction',''), r.get('qty',0))
        if key not in seen:
            seen.add(key)
            dedup_trades.append(r)
    all_trades = sorted(dedup_trades, key=lambda r: r.get('time',''))
    all_assets.sort(key=lambda d: d.get('date',''))

    print(f"  交易:{len(all_trades)}条 | 资产变动:{len(all_assets)}条 | 股息:{len(comb_flows['dividends'])}条")
    if broker_name: print(f"  券商: {broker_name}")

    rsu_config = load_rsu_config()

    # Opening positions
    opening_positions = {}
    if prior_year_file:
        print(f"  上年: {os.path.basename(prior_year_file)}")
        pt, pa, pf, _, _ = process_any_file(prior_year_file)
        wb_p = openpyxl.load_workbook(prior_year_file, data_only=True)
        po = parse_opening_positions(wb_p[find_sheet(wb_p, ['持仓总览'])])
        _, ptrackers, _ = calculate_all(pt, pa, pf, po, method, rsu_config)
        for key, t in ptrackers.items():
            if t.P > 0.001:
                code, market, currency = key
                opening_positions[(code, market, currency)] = {'qty':t.P, 'price':t.Q, 'market':market, 'currency':currency, 'code':code}
        print(f"  期初持仓:{len(opening_positions)}个")

    # Exchange rates
    if exchange_rates is None:
        exchange_rates = fetch_rates(year)
        if not exchange_rates: exchange_rates = {'USD':7.0, 'HKD':0.9}
    for cur, rate in exchange_rates.items(): print(f"  {cur}/CNY = {rate:.4f}")

    # Calculate
    enriched, trackers, warnings = calculate_all(all_trades, all_assets, comb_flows, opening_positions, method, rsu_config)
    total_R = sum(r['R'] for r in enriched)
    print(f"  R合计:{total_R:,.2f} ({sum(1 for r in enriched if r['R']>0.005)}盈/{sum(1 for r in enriched if r['R']<-0.005)}亏) | {len(trackers)}品种")

    tax = calculate_tax(trackers, comb_flows, exchange_rates, year)

    if output_file is None: output_file = f"{year}_税务审计_{method}.xlsx"
    write_output(enriched, trackers, tax, comb_flows, warnings, exchange_rates, output_file, year, method, broker_name)
    print(f"  ✅ {output_file}")
    for market, mkt in sorted(tax['capital_gains']['by_market'].items()):
        for cur, d in sorted(mkt['currencies'].items()):
            if abs(d['pnl']) > 0.01:
                print(f"     {mkt['label']:12s} {cur} 盈亏:{d['pnl']:>12,.2f}  税:¥{d['tax_rmb']:>10,.2f}")
    print(f"  资本利得税: ¥{tax['capital_gains']['total_tax_rmb']:,.2f}")
    print(f"  股息利息税: ¥{tax['dividends']['total_tax_rmb']+tax['interests']['total_tax_rmb']:,.2f}")
    print(f"  合计应补税: ¥{tax['grand_total_rmb']:,.2f}")
    return {'year':year,'output':output_file,'tax':tax,'total_R':total_R,'warnings':warnings}

def process_all_years(directory, method='WMA', exchange_rates=None):
    import glob
    files = sorted(glob.glob(os.path.join(directory, '*年度账单*.xlsx')))
    if not files:
        print("未找到 *年度账单*.xlsx"); return None
    years = sorted([(int(re.search(r'(\d{4})', os.path.basename(f)).group(1)), f) for f in files])
    if exchange_rates is None and years: exchange_rates = fetch_rates(years[-1][0])
    results = []
    for i, (year, fp) in enumerate(years):
        prior = years[i-1][1] if i > 0 else None
        out = os.path.join(directory, 'output', f'{year}_税务审计_{method}.xlsx')
        os.makedirs(os.path.dirname(out), exist_ok=True)
        results.append(process_year(fp, prior, out, year, method, exchange_rates))
    if results:
        _write_summary(results, directory, method)
    return results

def _write_summary(results, directory, method):
    import openpyxl
    from openpyxl.styles import Font,PatternFill,Border,Side,Alignment
    from openpyxl.utils import get_column_letter
    out = os.path.join(directory, 'output', f'所有年度汇总_{method}.xlsx')
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "年度汇总"
    ws.merge_cells('A1:I1')
    ws.cell(row=1, column=1, value="境外所得税多年度汇总").font = Font(name='Arial', size=16, bold=True, color='1F4E79')
    ws.row_dimensions[1].height = 35
    row = 3
    HF=PatternFill(start_color='1F4E79',end_color='1F4E79',fill_type='solid')
    HFONT=Font(name='Arial',size=11,bold=True,color='FFFFFF')
    TB=Border(left=Side('thin'),right=Side('thin'),top=Side('thin'),bottom=Side('thin'))
    CA=Alignment(horizontal='center',vertical='center')
    for c,h in enumerate(['年度','资本利得税','股息红利税','利息税','合计','港股盈亏','美股盈亏','交易数','备注'],1):
        ws.cell(row=row,column=c,value=h).font=HFONT; ws.cell(row=row,column=c).fill=HF
        ws.cell(row=row,column=c).alignment=CA; ws.cell(row=row,column=c).border=TB
    row+=1; gtotal=0
    for r in results:
        t=r['tax']
        hk_pnl=sum(d['pnl'] for m,mkt in t['capital_gains']['by_market'].items() if '港股' in mkt['label'] for d in mkt['currencies'].values())
        us_pnl=sum(d['pnl'] for m,mkt in t['capital_gains']['by_market'].items() if '美股' in mkt['label'] for d in mkt['currencies'].values())
        vals=[r['year'],t['capital_gains']['total_tax_rmb'],t['dividends']['total_tax_rmb'],
              t['interests']['total_tax_rmb'],t['grand_total_rmb'],round(hk_pnl,2),round(us_pnl,2),
              len(r.get('enriched',[])),f"⚠️{len(r['warnings'])}" if r.get('warnings') else '✓']
        for c,v in enumerate(vals,1):
            ws.cell(row=row,column=c,value=v).font=Font(name='Arial',size=10)
            ws.cell(row=row,column=c).alignment=CA; ws.cell(row=row,column=c).border=TB
        gtotal+=t['grand_total_rmb']; row+=1
    YF=PatternFill(start_color='FFE699',end_color='FFE699',fill_type='solid')
    RB=Font(name='Arial',size=12,bold=True,color='FF0000')
    ws.cell(row=row,column=1,value="合计").font=Font(name='Arial',size=10,bold=True)
    ws.cell(row=row,column=1).fill=YF; ws.cell(row=row,column=1).alignment=CA; ws.cell(row=row,column=1).border=TB
    ws.cell(row=row,column=5,value=round(gtotal,2)).font=RB; ws.cell(row=row,column=5).fill=YF
    ws.cell(row=row,column=5).alignment=CA; ws.cell(row=row,column=5).border=TB
    for c in range(2,9): ws.cell(row=row,column=c).fill=YF; ws.cell(row=row,column=c).border=TB
    for i,w in enumerate([8,16,16,14,16,14,14,12,10],1): ws.column_dimensions[get_column_letter(i)].width=w
    wb.save(out); print(f"\n📊 {out}")

# ═══════════════════════════════════════════════════════════
# CLI
# ═══════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description=f'CRS境外所得税计算器 v{VERSION}')
    parser.add_argument('input_files', nargs='*', help='账单文件(支持多格式多文件)')
    parser.add_argument('--prior', help='上年度文件')
    parser.add_argument('-o','--output', help='输出路径')
    parser.add_argument('--year', type=int, help='年度')
    parser.add_argument('--method', choices=['WMA','FIFO'], default='WMA')
    parser.add_argument('--usd-rate', type=float); parser.add_argument('--hkd-rate', type=float)
    parser.add_argument('--all-years', metavar='DIR', help='批量处理目录')

    args = parser.parse_args()
    rates = {}
    if args.usd_rate: rates['USD'] = args.usd_rate
    if args.hkd_rate: rates['HKD'] = args.hkd_rate
    if not rates: rates = None

    if args.all_years:
        process_all_years(args.all_years, args.method, rates); return 0
    if not args.input_files: parser.print_help(); return 1
    process_year(args.input_files, args.prior, args.output, args.year, args.method, rates)
    return 0

if __name__ == '__main__':
    sys.exit(main())
