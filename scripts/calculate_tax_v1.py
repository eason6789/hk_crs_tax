#!/usr/bin/env python3
"""
富途年度账单 CRS 税务计算器
计算境外股票/基金/期权转让所得的应税金额

用法:
  python calculate_tax.py <年度账单.xlsx> [上年度账单.xlsx] [-o 输出文件.xlsx]

功能:
  1. 解析富途年度账单 Excel 文件
  2. 按移动加权平均法计算每笔交易的 P(剩余股数)、Q(加权成本)、R(卖出盈亏)
  3. 分别按市场(US/HK)统计盈亏
  4. 统计股息、利息收入
  5. 计算应补税额(20%税率,含境外预提税抵免)
  6. 输出增强后的 Excel 文件
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
import sys
import os
import json
import argparse
from datetime import datetime, date
from collections import defaultdict
from copy import copy

# ═══════════════════════════════════════════════════════════════
# 汇率获取
# ═══════════════════════════════════════════════════════════════

def fetch_exchange_rates(year):
    """
    获取人民币汇率中间价(上年度最后一日)
    优先从网络获取,失败则提示用户手动输入

    数据来源: 中国外汇交易中心 / exchangerate-api
    """
    rates = {}

    # 尝试方法1: exchangerate-api (免费)
    try:
        import urllib.request
        target_date = f"{year}-12-31"
        for currency in ['USD', 'HKD']:
            url = f"https://api.exchangerate-api.com/v4/latest/{currency}"
            req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
            with urllib.request.urlopen(req, timeout=10) as resp:
                data = json.loads(resp.read().decode())
                rates[currency] = data['rates'].get('CNY', 0)
        if rates.get('USD') and rates.get('HKD'):
            print(f"✅ 已自动获取汇率: USD/CNY={rates['USD']:.4f}, HKD/CNY={rates['HKD']:.4f}")
            return rates
    except Exception as e:
        print(f"⚠️  自动获取汇率失败: {e}")

    # 尝试方法2: 手动输入
    print(f"\n请手动输入 {year}年12月31日 人民币汇率中间价:")
    for currency in ['USD', 'HKD']:
        while True:
            try:
                val = input(f"  {currency}/CNY: ").strip()
                rates[currency] = float(val)
                break
            except ValueError:
                print("  请输入有效数字")

    return rates


# ═══════════════════════════════════════════════════════════════
# 数据解析
# ═══════════════════════════════════════════════════════════════

def parse_trading_records(ws):
    """
    解析 证券-交易流水 sheet
    返回交易记录列表,按时间排序
    """
    records = []
    for row in range(2, ws.max_row + 1):
        time_str = ws.cell(row=row, column=1).value  # 成交时间
        account_name = str(ws.cell(row=row, column=2).value or '')
        account_num = str(ws.cell(row=row, column=3).value or '')
        category = str(ws.cell(row=row, column=4).value or '')     # 品类
        code = str(ws.cell(row=row, column=5).value or '')         # 代码名称
        market = str(ws.cell(row=row, column=6).value or '')       # 交易所/市场
        direction = str(ws.cell(row=row, column=7).value or '')    # 方向
        settle_date = str(ws.cell(row=row, column=8).value or '')  # 交收日期
        currency = str(ws.cell(row=row, column=9).value or '')     # 币种
        qty = float(ws.cell(row=row, column=10).value or 0)        # 数量/面值
        price = float(ws.cell(row=row, column=11).value or 0)      # 价格
        amount = float(ws.cell(row=row, column=12).value or 0)     # 成交金额
        fee = float(ws.cell(row=row, column=13).value or 0)        # 总费用
        net_amount = float(ws.cell(row=row, column=14).value or 0) # 变动金额

        if not time_str or not code:
            continue

        records.append({
            'time': str(time_str),
            'account_name': account_name,
            'account_num': account_num,
            'category': category,
            'code': code,
            'market': market,
            'direction': direction,
            'settle_date': settle_date,
            'currency': currency,
            'qty': qty,
            'price': price,
            'amount': amount,
            'fee': abs(fee),
            'net_amount': net_amount,
            'row': row,  # 原始行号
        })

    # 按时间排序
    records.sort(key=lambda r: r['time'])
    return records


def parse_asset_movements(ws):
    """
    解析 证券-资产进出 sheet
    获取 IPO 中签、RSU归属、结构化票据等非交易资产变动

    返回: dict of {code: [{date, type, qty, currency, remark}]}
    """
    movements = []
    for row in range(2, ws.max_row + 1):
        mv_date = ws.cell(row=row, column=1).value       # 日期
        account_name = str(ws.cell(row=row, column=2).value or '')
        account_num = str(ws.cell(row=row, column=3).value or '')
        category = str(ws.cell(row=row, column=4).value or '')
        code = str(ws.cell(row=row, column=5).value or '')
        market = str(ws.cell(row=row, column=6).value or '')
        direction = str(ws.cell(row=row, column=7).value or '')
        mv_type = str(ws.cell(row=row, column=8).value or '')   # 类型
        currency = str(ws.cell(row=row, column=9).value or '')
        qty = float(ws.cell(row=row, column=10).value or 0)      # 数量
        remark = str(ws.cell(row=row, column=11).value or '')    # 备注

        if not mv_date or not code:
            continue

        movements.append({
            'date': str(mv_date),
            'account_name': account_name,
            'account_num': account_num,
            'category': category,
            'code': code,
            'market': market,
            'direction': direction,
            'type': mv_type,
            'currency': currency,
            'qty': qty,
            'remark': remark,
        })

    return movements


def parse_fund_flows(ws):
    """
    解析 证券-资金进出 sheet
    获取股息、利息、IPO扣款、RSU税款等信息

    返回: {dividends: [...], interests: [...], other: [...]}
    """
    dividends = []
    interests = []
    ipo_payments = []
    rsu_tax = []
    note_coupons = []
    other = []

    for row in range(2, ws.max_row + 1):
        f_date = ws.cell(row=row, column=1).value
        account_name = str(ws.cell(row=row, column=2).value or '')
        account_num = str(ws.cell(row=row, column=3).value or '')
        f_type = str(ws.cell(row=row, column=4).value or '')    # 类型
        direction = str(ws.cell(row=row, column=5).value or '')  # 方向
        currency = str(ws.cell(row=row, column=6).value or '')
        amount = float(ws.cell(row=row, column=7).value or 0)
        remark = str(ws.cell(row=row, column=8).value or '')

        if not f_date:
            continue

        entry = {
            'date': str(f_date),
            'account_name': account_name,
            'account_num': account_num,
            'type': f_type,
            'direction': direction,
            'currency': currency,
            'amount': amount,
            'remark': remark,
        }

        # 分类识别
        type_lower = f_type.lower()
        remark_lower = remark.lower()

        # 股息识别
        if ('分红' in f_type or 'dividend' in remark_lower or
            ('公司行动' in f_type and 'dividend' in remark_lower)):
            dividends.append(entry)
        # 股息预扣税识别
        elif ('withholding tax' in remark_lower or
              ('公司行动' in f_type and 'tax' in remark_lower)):
            dividends.append(entry)
        # 利息识别
        elif '利息' in f_type:
            interests.append(entry)
        # Note/Fund 票息
        elif ('coupon' in remark_lower or '票息' in remark):
            note_coupons.append(entry)
        # IPO 扣款
        elif 'ipo' in f_type.lower() or 'ipo' in remark_lower:
            ipo_payments.append(entry)
        # RSU 相关
        elif ('长期激励' in f_type or '股票收益计划' in f_type or
              'rsu' in remark_lower or '长期激励' in remark):
            rsu_tax.append(entry)
        else:
            other.append(entry)

    return {
        'dividends': dividends,
        'interests': interests,
        'ipo_payments': ipo_payments,
        'rsu_tax': rsu_tax,
        'note_coupons': note_coupons,
        'other': other,
    }


def parse_opening_positions(ws):
    """
    解析 证券-持仓总览 sheet 的期初持仓
    返回: dict of {(code, market, currency): {qty, price, market_value}}
    """
    positions = {}
    for row in range(2, ws.max_row + 1):
        period_type = str(ws.cell(row=row, column=1).value or '')
        if '期初' not in period_type:
            continue

        code = str(ws.cell(row=row, column=6).value or '')
        market = str(ws.cell(row=row, column=7).value or '')
        currency = str(ws.cell(row=row, column=8).value or '')
        qty = float(ws.cell(row=row, column=9).value or 0)
        price = float(ws.cell(row=row, column=10).value or 0)

        key = (code, market, currency)
        if key in positions:
            # 合并同一代码的持仓（不同账户）
            positions[key]['qty'] += qty
            # 加权平均价格
            old = positions[key]
            total_qty = old['qty'] + qty
            old['price'] = (old['price'] * old['qty'] + price * qty) / total_qty if total_qty > 0 else 0
            old['qty'] = total_qty
        else:
            positions[key] = {'qty': qty, 'price': price, 'market': market, 'currency': currency, 'code': code}

    return positions


# ═══════════════════════════════════════════════════════════════
# 核心计算逻辑 - 移动加权平均法
# ═══════════════════════════════════════════════════════════════

class PositionTracker:
    """单个证券的持仓跟踪器"""

    def __init__(self, code, market, currency):
        self.code = code
        self.market = market
        self.currency = currency
        self.P = 0.0          # 剩余股数
        self.Q = 0.0          # 加权平均成本(每股)
        self.total_cost = 0.0 # 总持仓成本(含买入手续费)
        self.realized_pnl = 0.0  # 累计已实现盈亏

        # 额外记录
        self.total_buy_amount = 0.0   # 累计买入金额(不含费)
        self.total_buy_fee = 0.0      # 累计买入手续费
        self.total_sell_amount = 0.0  # 累计卖出金额
        self.total_sell_fee = 0.0     # 累计卖出手续费

    def add_position(self, qty, cost_per_share):
        """添加期初持仓（从上年度结转）"""
        if qty <= 0:
            return
        self.P = qty
        self.Q = cost_per_share
        self.total_cost = qty * cost_per_share

    def buy(self, qty, price, fee=0.0):
        """
        买入/申购操作
        qty: 买入股数(正数)
        price: 成交价格
        fee: 买入手续费(计入成本)
        """
        if qty <= 0:
            return

        buy_amount = qty * price
        buy_cost = buy_amount + fee  # 买入手续费计入成本

        self.total_buy_amount += buy_amount
        self.total_buy_fee += fee

        if self.P > 0:
            # 重新计算加权平均成本
            self.total_cost += buy_cost
            self.P += qty
            self.Q = self.total_cost / self.P
        else:
            # 新开仓
            self.P = qty
            self.total_cost = buy_cost
            self.Q = self.total_cost / self.P

    def sell(self, qty, price, fee=0.0):
        """
        卖出/赎回操作
        qty: 卖出股数(正数)
        price: 成交价格
        fee: 卖出手续费(从收益中扣除)

        返回: realized_pnl (这笔卖出的已实现盈亏)
        """
        if qty <= 0 or self.P <= 0:
            return 0.0

        sell_qty = min(qty, self.P)  # 不能卖超过持仓

        # 卖出盈亏 = (卖出价 - 加权成本) × 股数 - 卖出手续费
        gross_pnl = (price - self.Q) * sell_qty
        realized = gross_pnl - fee

        self.total_sell_amount += sell_qty * price
        self.total_sell_fee += fee
        self.realized_pnl += realized

        # 更新持仓
        self.P -= sell_qty
        self.total_cost = self.P * self.Q  # 剩余持仓成本

        # 如果清仓，重置
        if abs(self.P) < 0.000001:  # 浮点精度
            self.P = 0.0
            self.Q = 0.0
            self.total_cost = 0.0

        return realized

    def get_summary(self):
        """获取持仓摘要"""
        return {
            'code': self.code,
            'market': self.market,
            'currency': self.currency,
            'P': round(self.P, 6),
            'Q': round(self.Q, 6),
            'total_cost': round(self.total_cost, 4),
            'realized_pnl': round(self.realized_pnl, 4),
            'total_buy_amount': round(self.total_buy_amount, 4),
            'total_buy_fee': round(self.total_buy_fee, 4),
            'total_sell_amount': round(self.total_sell_amount, 4),
            'total_sell_fee': round(self.total_sell_fee, 4),
            'unrealized_value': round(self.P * self.Q, 4),
        }


def calculate_all_positions(trading_records, asset_movements, fund_flows, opening_positions):
    """
    主计算函数：遍历所有交易记录，计算 P, Q, R

    返回:
      - enriched_records: 添加了 P, Q, R 列的交易记录
      - trackers: 每个证券的持仓跟踪器
      - summary: 按市场/币种汇总
    """
    # 初始化跟踪器
    trackers = {}  # key: (code, market, currency)

    def get_tracker(code, market, currency):
        key = (code, market, currency)
        if key not in trackers:
            trackers[key] = PositionTracker(code, market, currency)
        return trackers[key]

    # 1. 先加载期初持仓
    for (code, market, currency), pos in opening_positions.items():
        tracker = get_tracker(code, market, currency)
        tracker.add_position(pos['qty'], pos['price'])

    # 2. 处理资产进出（IPO、RSU 等非交易获得）
    # 注意：账户升级类型的In记录通常只是旧账户持仓迁移，这些股票已在期初持仓中
    # 只有真正的新增（IPO中签、RSU归属等）才需要加入
    asset_entries = []
    for mv in asset_movements:
        if mv['direction'] == 'In':
            mv_type = mv['type']
            code = mv['code']
            market = mv['market']
            currency = mv['currency']

            # 账户升级/账户迁移: 这些股票已为期初持仓，跳过避免重复
            if '账户升级' in mv_type or '账户迁移' in mv_type or '升级' in mv_type:
                continue

            # 查找对应的成本信息
            cost_info = find_asset_cost(mv, fund_flows, trading_records)
            asset_entries.append({
                'time': mv['date'],
                'code': code,
                'market': market,
                'currency': currency,
                'qty': mv['qty'],
                'cost_per_share': cost_info['price'],
                'fee': cost_info.get('fee', 0),
                'type': mv_type,
                'remark': mv['remark'],
                'needs_manual_price': cost_info.get('needs_manual_price', False),
            })

    # 3. 合并交易记录和资产变动，按时间排序
    all_events = []

    for rec in trading_records:
        all_events.append({
            'type': 'trade',
            'time': rec['time'],
            'data': rec,
        })

    for entry in asset_entries:
        all_events.append({
            'type': 'asset_in',
            'time': entry['time'],
            'data': entry,
        })

    all_events.sort(key=lambda e: e['time'])

    # 4. 逐笔处理
    enriched_records = []

    for event in all_events:
        if event['type'] == 'asset_in':
            data = event['data']
            tracker = get_tracker(data['code'], data['market'], data['currency'])
            if data['qty'] > 0:
                cost = data.get('cost_per_share', 0)
                if cost <= 0:
                    # RSU/IPO等非交易获得的资产，成本未知时仍计入持仓
                    # 但标记为需要手动确认成本
                    needs_manual = data.get('needs_manual_price', True)
                    label = data.get('type', '未知')
                    print(f"   ⚠️  {data['code']} ({label}): +{data['qty']:,.0f}股, 成本未知，暂按0计入")
                    print(f"       这会导致该证券的Q(加权成本)偏低、R(卖出盈亏)偏高")
                    print(f"       请手动查找 {label} 的获取成本后重新计算")
                tracker.buy(data['qty'], max(cost, 0.0001), data.get('fee', 0))
            continue

        # 交易记录
        rec = event['data']
        tracker = get_tracker(rec['code'], rec['market'], rec['currency'])

        direction = rec['direction']
        qty = abs(rec['qty'])
        price = rec['price']
        fee = rec['fee']

        P_before = tracker.P
        Q_before = tracker.Q

        realized = 0.0

        # 判断买卖方向
        is_buy = direction in ('买入开仓', '买入平仓', '申购', '买入')
        is_sell = direction in ('卖出平仓', '卖出开仓', '赎回', '卖出')

        if is_buy and rec['amount'] < 0:
            # 买入: 成交金额为负(支出)
            tracker.buy(qty, price, fee)
        elif is_buy and rec['amount'] > 0:
            # 有些情况下成交金额为正(如基金赎回时为正收入)
            # 按方向判断: 申购=买入
            tracker.buy(qty, price, fee)
        elif is_sell:
            # 卖出: 计算盈亏
            realized = tracker.sell(qty, price, fee)

        # 记录增强数据
        enriched = dict(rec)
        enriched['P_before'] = round(P_before, 6)
        enriched['Q_before'] = round(Q_before, 6)
        enriched['P_after'] = round(tracker.P, 6)
        enriched['Q_after'] = round(tracker.Q, 6)
        enriched['R'] = round(realized, 4)  # 卖出盈亏

        # 对于买入，R 列显示相对于当前成本的差异(便于验证)
        if is_buy and not is_sell:
            enriched['R'] = 0.0

        enriched_records.append(enriched)

    return enriched_records, trackers


def find_asset_cost(mv, fund_flows, trading_records=None):
    """
    查找非交易获得资产（IPO、RSU等）的成本信息

    优先级:
    1. 查找同日期/相邻日期的 IPO 扣款记录
    2. 对于 RSU，从税款记录反推/附近交易价格估算
    3. 从备注中提取价格信息
    """
    result = {'price': 0.0, 'fee': 0.0}

    mv_type = mv['type']
    remark = mv['remark']
    code = mv['code']
    mv_date = mv['date']
    mv_qty = mv['qty']

    # IPO 中签: 在资金进出中查找对应 IPO 扣款
    if 'ipo' in mv_type.lower() or 'ipo' in remark.lower():
        for payment in fund_flows.get('ipo_payments', []):
            if payment['direction'] == 'Out' and abs(date_diff(payment['date'], mv_date)) <= 3:
                result['price'] = abs(payment['amount']) / mv_qty if mv_qty > 0 else 0
                return result

    # RSU/长期激励/股票收益计划
    if any(kw in mv_type for kw in ['长期激励', '股票收益计划', '激励']) or 'rsu' in remark.lower():
        # 策略1: 从税款记录估算
        for rsu_entry in fund_flows.get('rsu_tax', []):
            if rsu_entry['direction'] == 'Out' and abs(date_diff(rsu_entry['date'], mv_date)) <= 3:
                # RSU税款是 vesting 时按工资薪金预扣的
                # 中国RSU预扣税率约20-45%，用中间值22%作为粗略估算
                tax_amount = abs(rsu_entry['amount'])
                if tax_amount > 0 and mv_qty > 0:
                    # 保守估算：假设实际税率较低 → FMV较高 → 成本偏高 → 税负偏低（安全边际）
                    est_fmv_per_share = tax_amount / (0.20 * mv_qty)
                    # 合理性检查：如果估算值明显不合理（太低），尝试其他税率
                    if est_fmv_per_share < 1.0:
                        est_fmv_per_share = tax_amount / (0.10 * mv_qty)
                    result['price'] = round(est_fmv_per_share, 4)
                    result['estimated_from_tax'] = True
                    return result

        # 策略2: 从附近交易价格估算
        if trading_records:
            nearby_prices = []
            for rec in trading_records:
                if rec['code'] == code and rec['market'] == mv['market']:
                    try:
                        if abs(date_diff(rec['time'][:8], mv_date)) <= 30:
                            nearby_prices.append(rec['price'])
                    except:
                        pass
            if nearby_prices:
                result['price'] = round(sum(nearby_prices) / len(nearby_prices), 4)
                result['estimated_from_market'] = True
                return result

        # 策略3: 无法自动确定，标记为需要手动输入
        result['price'] = 0.0
        result['needs_manual_price'] = True
        return result

    # 结构化票据
    if '结构化票据' in mv_type or 'structured' in remark.lower() or 'fcn' in remark.lower():
        # 结构化票据的票面金额就是成本
        result['price'] = 1.0  # 按面值
        result['needs_manual_price'] = True

    return result


def date_diff(d1, d2):
    """计算两个日期字符串的天数差"""
    try:
        fmt = '%Y%m%d'
        dt1 = datetime.strptime(d1.replace('-', '')[:8], fmt)
        dt2 = datetime.strptime(d2.replace('-', '')[:8], fmt)
        return abs((dt1 - dt2).days)
    except:
        return 999


# ═══════════════════════════════════════════════════════════════
# 税务计算
# ═══════════════════════════════════════════════════════════════

def calculate_tax_summary(trackers, fund_flows, exchange_rates):
    """
    计算税务汇总

    返回:
      - capital_gains_tax: 资本利得税(按市场/币种)
      - dividend_tax: 股息税(含预提税抵免)
      - interest_tax: 利息税
      - total_tax_rmb: 总应补税额(RMB)
    """
    # 1. 资本利得汇总(按市场)
    market_pnl = defaultdict(lambda: defaultdict(float))  # {market: {currency: pnl}}

    for key, tracker in trackers.items():
        code, market, currency = key
        if tracker.realized_pnl != 0:
            market_pnl[market][currency] += tracker.realized_pnl

    # 2. 股息汇总
    dividend_summary = defaultdict(lambda: {'gross': 0.0, 'tax_withheld': 0.0, 'net': 0.0, 'currency': ''})

    for div in fund_flows.get('dividends', []):
        remark = div['remark'].lower()
        if div['direction'] == 'In' and ('dividend' in remark or 'i/d' in remark.lower()):
            # 股息收入
            key = f"{div['currency']}_DIV"
            dividend_summary[key]['gross'] += div['amount']
            dividend_summary[key]['currency'] = div['currency']
        elif div['direction'] == 'Out' and ('withholding tax' in remark or 'tax' in remark):
            # 预扣税
            key = f"{div['currency']}_DIV"
            dividend_summary[key]['tax_withheld'] += abs(div['amount'])
            dividend_summary[key]['currency'] = div['currency']

    for key in dividend_summary:
        dividend_summary[key]['net'] = dividend_summary[key]['gross'] - dividend_summary[key]['tax_withheld']

    # 3. 利息汇总
    interest_total = defaultdict(float)
    for interest in fund_flows.get('interests', []):
        # 利息扣除是支出（负值），只有利息收入需要计税
        if interest['direction'] == 'In':
            interest_total[interest['currency']] += interest['amount']

    # 4. Note/Fund 票息
    for coupon in fund_flows.get('note_coupons', []):
        if coupon['direction'] == 'In':
            interest_total[coupon['currency']] += coupon['amount']

    # 5. 折算为 RMB 计算税额
    capital_gains_tax_rmb = 0.0
    capital_gains_detail = {}

    for market, currencies in market_pnl.items():
        market_tax = 0.0
        detail = {}
        for currency, pnl in currencies.items():
            rate = exchange_rates.get(currency, 1.0)
            pnl_rmb = pnl * rate
            tax_rmb = max(0, pnl_rmb * 0.20)  # 盈利才计税，亏损不退税
            market_tax += tax_rmb
            detail[currency] = {
                'pnl': round(pnl, 2),
                'rate': rate,
                'pnl_rmb': round(pnl_rmb, 2),
                'tax_rmb': round(tax_rmb, 2),
            }
        capital_gains_tax_rmb += market_tax
        capital_gains_detail[market] = {
            'currencies': detail,
            'subtotal_tax_rmb': round(market_tax, 2),
        }

    # 股息税计算
    dividend_tax_rmb = 0.0
    dividend_detail = {}
    for key, div_data in dividend_summary.items():
        currency = div_data['currency']
        rate = exchange_rates.get(currency, 1.0)
        gross_rmb = div_data['gross'] * rate
        tax_should = gross_rmb * 0.20  # 应缴20%
        tax_withheld_rmb = div_data['tax_withheld'] * rate
        tax_owed = max(0, tax_should - tax_withheld_rmb)  # 境外预扣税可抵免

        dividend_tax_rmb += tax_owed
        dividend_detail[key] = {
            'gross': round(div_data['gross'], 2),
            'gross_rmb': round(gross_rmb, 2),
            'tax_withheld': round(div_data['tax_withheld'], 2),
            'tax_withheld_rmb': round(tax_withheld_rmb, 2),
            'tax_should_rmb': round(tax_should, 2),
            'tax_owed_rmb': round(tax_owed, 2),
            'currency': currency,
        }

    # 利息税
    interest_tax_rmb = 0.0
    interest_detail = {}
    for currency, total in interest_total.items():
        rate = exchange_rates.get(currency, 1.0)
        total_rmb = total * rate
        tax = total_rmb * 0.20
        interest_tax_rmb += tax
        interest_detail[currency] = {
            'total': round(total, 2),
            'total_rmb': round(total_rmb, 2),
            'tax_rmb': round(tax, 2),
        }

    return {
        'capital_gains': {
            'detail': capital_gains_detail,
            'subtotal_rmb': round(capital_gains_tax_rmb, 2),
        },
        'dividends': {
            'detail': dividend_detail,
            'subtotal_rmb': round(dividend_tax_rmb, 2),
        },
        'interests': {
            'detail': interest_detail,
            'subtotal_rmb': round(interest_tax_rmb, 2),
        },
        'total_tax_rmb': round(capital_gains_tax_rmb + dividend_tax_rmb + interest_tax_rmb, 2),
    }


# ═══════════════════════════════════════════════════════════════
# 输出生成
# ═══════════════════════════════════════════════════════════════

def write_output_excel(enriched_records, trackers, tax_summary, exchange_rates, fund_flows, output_path, year):
    """生成输出 Excel 文件"""
    wb = openpyxl.Workbook()

    # ── Sheet 1: 交易流水(增强版) ──
    ws1 = wb.active
    ws1.title = "交易流水-含PQR"

    # 表头
    headers = [
        '成交时间', '账户名称', '账户号码', '品类', '代码名称', '交易所/市场',
        '方向', '交收日期', '币种', '数量/面值', '价格', '成交金额', '总费用',
        '变动金额',
        'P(交易前持仓)', 'Q(交易前成本)', 'P(交易后持仓)', 'Q(交易后成本)',
        'R(卖出盈亏)'
    ]

    # 样式
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font_white = Font(bold=True, size=11, color='FFFFFF')
    pnl_green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    pnl_red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # 数据行
    for i, rec in enumerate(enriched_records):
        row = i + 2
        values = [
            rec['time'], rec['account_name'], rec['account_num'],
            rec['category'], rec['code'], rec['market'],
            rec['direction'], rec['settle_date'], rec['currency'],
            rec['qty'], rec['price'], rec['amount'], rec['fee'],
            rec['net_amount'],
            rec['P_before'], rec['Q_before'],
            rec['P_after'], rec['Q_after'],
            rec['R'],
        ]
        for col, val in enumerate(values, 1):
            cell = ws1.cell(row=row, column=col, value=val)

        # R 列着色: 盈利绿色，亏损红色
        r_cell = ws1.cell(row=row, column=19)
        if rec['R'] > 0:
            r_cell.fill = pnl_green_fill
        elif rec['R'] < 0:
            r_cell.fill = pnl_red_fill

    # 列宽调整
    col_widths = [20, 30, 18, 8, 12, 10, 10, 10, 8, 14, 10, 14, 10, 12, 14, 14, 14, 14, 14]
    for i, w in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # 冻结首行
    ws1.freeze_panes = 'A2'

    # ── Sheet 2: 持仓汇总 ──
    ws2 = wb.create_sheet("持仓汇总")
    pos_headers = [
        '代码名称', '交易所/市场', '币种', '剩余股数(P)', '加权成本(Q)',
        '总持仓成本', '累计已实现盈亏', '累计买入金额', '累计买入手续费',
        '累计卖出金额', '累计卖出手续费'
    ]

    for col, header in enumerate(pos_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for i, (key, tracker) in enumerate(sorted(trackers.items())):
        row = i + 2
        code, market, currency = key
        s = tracker.get_summary()
        values = [code, market, currency, s['P'], s['Q'], s['total_cost'],
                  s['realized_pnl'], s['total_buy_amount'], s['total_buy_fee'],
                  s['total_sell_amount'], s['total_sell_fee']]
        for col, val in enumerate(values, 1):
            ws2.cell(row=row, column=col, value=val)

    ws2.freeze_panes = 'A2'

    # ── Sheet 3: 税务汇总 ──
    ws3 = wb.create_sheet("税务汇总")

    row = 1
    # 标题
    ws3.cell(row=row, column=1, value=f"{year}年度 境外所得税计算汇总").font = Font(bold=True, size=14)
    row += 2

    # 汇率信息
    ws3.cell(row=row, column=1, value="使用汇率(人民币中间价):").font = Font(bold=True)
    for currency, rate in exchange_rates.items():
        row += 1
        ws3.cell(row=row, column=1, value=f"  {currency}/CNY = {rate:.4f}")
    row += 2

    # 一、资本利得税
    ws3.cell(row=row, column=1, value="一、财产转让所得(资本利得)").font = Font(bold=True, size=12)
    row += 1

    cg_headers = ['市场', '币种', '年度净盈亏', '汇率', '盈亏(RMB)', '应纳税额(RMB)']
    for col, h in enumerate(cg_headers, 1):
        cell = ws3.cell(row=row, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
    row += 1

    cg = tax_summary['capital_gains']
    for market, mkt_data in cg['detail'].items():
        for currency, cur_data in mkt_data['currencies'].items():
            ws3.cell(row=row, column=1, value=market)
            ws3.cell(row=row, column=2, value=currency)
            ws3.cell(row=row, column=3, value=cur_data['pnl'])
            ws3.cell(row=row, column=4, value=cur_data['rate'])
            ws3.cell(row=row, column=5, value=cur_data['pnl_rmb'])
            ws3.cell(row=row, column=6, value=cur_data['tax_rmb'])
            if cur_data['tax_rmb'] > 0:
                ws3.cell(row=row, column=6).font = Font(bold=True, color='FF0000')
            row += 1

    ws3.cell(row=row, column=1, value="资本利得税小计").font = Font(bold=True)
    ws3.cell(row=row, column=6, value=cg['subtotal_rmb']).font = Font(bold=True, color='FF0000')
    row += 2

    # 二、股息红利税
    ws3.cell(row=row, column=1, value="二、利息、股息、红利所得").font = Font(bold=True, size=12)
    row += 1

    div_headers = ['类型', '币种', '股息总额', '股息总额(RMB)', '境外预扣税', '预扣税(RMB)', '应缴税(RMB)', '应补税(RMB)']
    for col, h in enumerate(div_headers, 1):
        cell = ws3.cell(row=row, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
    row += 1

    for key, div_data in tax_summary['dividends']['detail'].items():
        ws3.cell(row=row, column=1, value='股息')
        ws3.cell(row=row, column=2, value=div_data['currency'])
        ws3.cell(row=row, column=3, value=div_data['gross'])
        ws3.cell(row=row, column=4, value=div_data['gross_rmb'])
        ws3.cell(row=row, column=5, value=div_data['tax_withheld'])
        ws3.cell(row=row, column=6, value=div_data['tax_withheld_rmb'])
        ws3.cell(row=row, column=7, value=div_data['tax_should_rmb'])
        ws3.cell(row=row, column=8, value=div_data['tax_owed_rmb'])
        row += 1

    # 利息
    for currency, int_data in tax_summary['interests']['detail'].items():
        ws3.cell(row=row, column=1, value='利息/票息')
        ws3.cell(row=row, column=2, value=currency)
        ws3.cell(row=row, column=3, value=int_data['total'])
        ws3.cell(row=row, column=4, value=int_data['total_rmb'])
        ws3.cell(row=row, column=5, value=0)
        ws3.cell(row=row, column=6, value=0)
        ws3.cell(row=row, column=7, value=int_data['tax_rmb'])
        ws3.cell(row=row, column=8, value=int_data['tax_rmb'])
        row += 1

    div_sub = tax_summary['dividends']['subtotal_rmb']
    int_sub = tax_summary['interests']['subtotal_rmb']
    ws3.cell(row=row, column=1, value="股息利息税小计").font = Font(bold=True)
    ws3.cell(row=row, column=8, value=round(div_sub + int_sub, 2)).font = Font(bold=True, color='FF0000')
    row += 2

    # 三、合计
    total = tax_summary['total_tax_rmb']
    ws3.cell(row=row, column=1, value="三、合计应补税额").font = Font(bold=True, size=14)
    row += 1
    ws3.cell(row=row, column=1, value=f"应补缴个人所得税总额: ¥{total:,.2f}").font = Font(bold=True, size=14, color='FF0000')
    row += 1
    ws3.cell(row=row, column=1, value=f"(不含滞纳金，请于次年6月30日前完成申报)").font = Font(italic=True)
    row += 2

    # 滞纳金提醒
    ws3.cell(row=row, column=1, value="⚠️ 重要提醒:").font = Font(bold=True, color='FF0000')
    row += 1
    reminders = [
        "• 申报时间: 次年3月1日至6月30日，逾期按日加收万分之五滞纳金",
        "• 浮盈不征税: 仅对实际卖出实现的盈利计税，未卖出持仓无需缴税",
        "• 跨年度亏损不能结转抵扣",
        "• 美股股息默认预扣10%(W8BEN)，可申请税收抵免，仅需补缴差额",
        "• 港股通过港股通持股的股息已预扣10%，同样可申请抵免",
        "• RSU归属时已按工资薪金计税(由雇主扣缴)，卖出时按RSU归属日市价为成本",
        "• 建议保留完整交易记录和完税证明备查",
    ]
    for reminder in reminders:
        ws3.cell(row=row, column=1, value=reminder)
        row += 1

    # 列宽
    for i, w in enumerate([35, 10, 15, 12, 15, 15, 15, 15], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 4: 股息明细 ──
    ws4 = wb.create_sheet("股息明细")
    div_detail_headers = ['日期', '账户', '类型', '方向', '币种', '金额', '备注']
    for col, h in enumerate(div_detail_headers, 1):
        cell = ws4.cell(row=1, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill

    row = 2
    for div in fund_flows.get('dividends', []):
        ws4.cell(row=row, column=1, value=div['date'])
        ws4.cell(row=row, column=2, value=div['account_name'])
        ws4.cell(row=row, column=3, value=div['type'])
        ws4.cell(row=row, column=4, value=div['direction'])
        ws4.cell(row=row, column=5, value=div['currency'])
        ws4.cell(row=row, column=6, value=div['amount'])
        ws4.cell(row=row, column=7, value=div['remark'])
        row += 1

    # 保存
    wb.save(output_path)
    print(f"\n✅ 输出文件已保存: {output_path}")


# ═══════════════════════════════════════════════════════════════
# 主程序
# ═══════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description='富途年度账单 CRS 税务计算器',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
使用示例:
  python calculate_tax.py 2025_年度账单_11267892.xlsx
  python calculate_tax.py 2025_年度账单_11267892.xlsx 2024_年度账单_11267892.xlsx
  python calculate_tax.py 2025_年度账单_11267892.xlsx -o 2025_税务计算结果.xlsx
        """
    )
    parser.add_argument('input_file', help='当年度的富途年度账单 Excel 文件')
    parser.add_argument('prior_year_file', nargs='?', help='上年度富途年度账单(用于期初持仓)')
    parser.add_argument('-o', '--output', default=None, help='输出文件路径(默认自动生成)')
    parser.add_argument('--year', type=int, help='纳税年度(默认从文件名提取)')
    parser.add_argument('--usd-rate', type=float, help='手动指定 USD/CNY 汇率')
    parser.add_argument('--hkd-rate', type=float, help='手动指定 HKD/CNY 汇率')

    args = parser.parse_args()

    # 确定年度
    if args.year:
        year = args.year
    else:
        # 从文件名提取
        import re
        match = re.search(r'(\d{4})', os.path.basename(args.input_file))
        year = int(match.group(1)) if match else datetime.now().year

    tax_year_end = year  # 纳税年度最后一日所在年份

    print(f"{'='*60}")
    print(f"  富途 CRS 税务计算器 - {year}年度")
    print(f"{'='*60}")

    # 1. 读取主要文件
    print(f"\n📂 读取文件: {args.input_file}")
    wb = openpyxl.load_workbook(args.input_file, data_only=True)

    # 解析交易流水
    print("📊 解析交易流水...")
    ws_trading = wb['证券-交易流水']
    trading_records = parse_trading_records(ws_trading)
    print(f"   共 {len(trading_records)} 条交易记录")

    # 解析资产进出
    print("📊 解析资产进出...")
    ws_assets = wb['证券-资产进出']
    asset_movements = parse_asset_movements(ws_assets)
    print(f"   共 {len(asset_movements)} 条资产变动")

    # 解析资金进出
    print("📊 解析资金进出...")
    ws_funds = wb['证券-资金进出']
    fund_flows = parse_fund_flows(ws_funds)
    print(f"   股息相关: {len(fund_flows['dividends'])} 条")
    print(f"   利息相关: {len(fund_flows['interests'])} 条")
    print(f"   IPO扣款: {len(fund_flows['ipo_payments'])} 条")
    print(f"   票息/其他: {len(fund_flows['note_coupons'])} 条")

    # 2. 处理期初持仓和跨年度数据
    # 关键原则：期初持仓的「价格」是市价而非成本价，不能用市价当成本。
    # 正确做法：处理上年度全部交易记录来建立准确的成本基准。
    opening_positions = {}

    if args.prior_year_file:
        print(f"\n📂 读取上年度文件: {args.prior_year_file}")
        wb_prior = openpyxl.load_workbook(args.prior_year_file, data_only=True)

        # 解析上年度交易记录
        print("📊 解析上年度交易流水(用于成本基准)...")
        ws_prior_trade = wb_prior['证券-交易流水']
        prior_trading = parse_trading_records(ws_prior_trade)

        # 解析上年度资产进出
        ws_prior_assets = wb_prior['证券-资产进出']
        prior_assets = parse_asset_movements(ws_prior_assets)

        # 解析上年度资金进出
        ws_prior_funds = wb_prior['证券-资金进出']
        prior_flows = parse_fund_flows(ws_prior_funds)

        print(f"   上年度交易: {len(prior_trading)} 条, 资产变动: {len(prior_assets)} 条")

        # 读取上年度自己的期初持仓（用于上年度计算的起点）
        ws_prior_positions = wb_prior['证券-持仓总览']
        prior_opening = parse_opening_positions(ws_prior_positions)
        print(f"   上年度期初持仓: {len(prior_opening)} 个")

        # 先跑一遍上年度计算，建立成本基准
        print("🔢 计算上年度成本基准...")
        prior_enriched, prior_trackers = calculate_all_positions(
            prior_trading, prior_assets, prior_flows, prior_opening
        )

        # 从上年度 trackers 提取期末成本作为本年度期初成本
        for key, tracker in prior_trackers.items():
            if tracker.P > 0:
                code, market, currency = key
                opening_positions[(code, market, currency)] = {
                    'qty': tracker.P,
                    'price': tracker.Q,  # 使用上年度计算的加权成本！
                    'market': market,
                    'currency': currency,
                    'code': code,
                }

        print(f"   结转期初持仓(含成本): {len(opening_positions)} 个")
        for (code, market, currency), pos in sorted(opening_positions.items()):
            print(f"     {code} ({market}) {currency}: {pos['qty']:,.0f}股 @ {pos['price']:.4f} 成本")

    else:
        # 无上年度文件：从当年文件期初读取（市价为近似成本）
        if '证券-持仓总览' in wb.sheetnames:
            ws_opening = wb['证券-持仓总览']
            opening_positions = parse_opening_positions(ws_opening)
            if opening_positions:
                print(f"\n📊 从当年文件中读取到期初持仓: {len(opening_positions)} 个")
                print("⚠️  ⚠️  ⚠️  重要警告 ⚠️  ⚠️  ⚠️")
                print("   未提供上年度文件，期初持仓成本使用期末市价作为近似值。")
                print("   这可能导致税务计算不准确（市价 ≠ 实际买入成本）！")
                print("   强烈建议同时提供上年度文件以获得准确的成本基准。")
                print("   如果这些持仓是多年前买入的，误差可能很大。")

    # 3. 获取汇率
    if args.usd_rate and args.hkd_rate:
        exchange_rates = {'USD': args.usd_rate, 'HKD': args.hkd_rate}
        print(f"\n💰 使用手动汇率: USD/CNY={args.usd_rate}, HKD/CNY={args.hkd_rate}")
    else:
        print(f"\n💰 获取汇率...")
        exchange_rates = fetch_exchange_rates(tax_year_end)

    # 4. 核心计算 - 仅处理本年度交易
    print("\n🔢 开始计算本年度 P/Q/R...")
    enriched_records, trackers = calculate_all_positions(
        trading_records, asset_movements, fund_flows, opening_positions
    )

    # 统计（区分币种）
    r_by_currency = defaultdict(float)
    for r in enriched_records:
        r_by_currency[r['currency']] += r['R']
    total_R = sum(r['R'] for r in enriched_records)
    profit_count = sum(1 for r in enriched_records if r['R'] > 0)
    loss_count = sum(1 for r in enriched_records if r['R'] < 0)
    print(f"   卖出盈亏总和(R列): {total_R:,.2f}")
    if len(r_by_currency) > 1:
        for cur, val in sorted(r_by_currency.items()):
            print(f"      {cur}: {val:,.2f}")
    print(f"   盈利交易: {profit_count} 笔, 亏损交易: {loss_count} 笔")
    print(f"   跟踪证券数: {len(trackers)}")

    # 5. 税务计算
    print("\n🧮 计算税务...")
    tax_summary = calculate_tax_summary(trackers, fund_flows, exchange_rates)

    print(f"\n{'='*60}")
    print(f"  📋 税务计算结果")
    print(f"{'='*60}")
    print(f"  资本利得税: ¥{tax_summary['capital_gains']['subtotal_rmb']:,.2f}")
    print(f"  股息红利税: ¥{tax_summary['dividends']['subtotal_rmb']:,.2f}")
    print(f"  利息收入税: ¥{tax_summary['interests']['subtotal_rmb']:,.2f}")
    print(f"  {'─'*30}")
    print(f"  合计应补税: ¥{tax_summary['total_tax_rmb']:,.2f}")
    print(f"{'='*60}")

    # 6. 输出
    if args.output:
        output_path = args.output
    else:
        base = os.path.splitext(os.path.basename(args.input_file))[0]
        output_path = f"{base}_税务计算.xlsx"

    print(f"\n💾 生成输出文件...")
    write_output_excel(enriched_records, trackers, tax_summary, exchange_rates, fund_flows, output_path, year)

    print(f"\n🎉 计算完成!")
    return 0


if __name__ == '__main__':
    sys.exit(main())
