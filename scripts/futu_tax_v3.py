#!/usr/bin/env python3
"""
╔══════════════════════════════════════════════════════════════════════╗
║    富途/券商 CRS 境外所得税计算器 v3.0                              ║
║    Universal Offshore Income Tax Calculator                         ║
║                                                                    ║
║  支持: 年度账单Excel | 月结单 | 日结单 | PDF (需crs-report-generator)║
║  品种: 股票/基金/期权/期货/债券/结构化票据                           ║
║  方法: 移动加权平均(WMA) / 先进先出(FIFO)                           ║
╚══════════════════════════════════════════════════════════════════════╝
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys, os, json, re, argparse
from datetime import datetime
from collections import defaultdict
import urllib.request, ssl

VERSION = "3.0.0"
TAX_RATE = 0.20
RSU_CONFIG_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'rsu_config.json')

# ── Market classification ──
MARKET_MAP = {'US':'🇺🇸 美股','SEHK':'🇭🇰 港股','FD':'🌐 基金','-':'🌐 其他','FTSN':'🌐 票据'}
# ── Product category mapping ──
CATEGORY_MAP = {'证券':'stock','基金':'fund','期权':'option','期货':'future','债券':'bond','结构化票据':'note'}

# ── Styles ──
HDR_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
SUB_FILL = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
GRN_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
RED_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
YEL_FILL = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')
HDR_FONT = Font(name='Arial', size=11, bold=True, color='FFFFFF')
TITLE_FONT = Font(name='Arial', size=16, bold=True, color='1F4E79')
NORM_FONT = Font(name='Arial', size=10)
BOLD_FONT = Font(name='Arial', size=10, bold=True)
RED_BOLD = Font(name='Arial', size=12, bold=True, color='FF0000')
THIN_BORDER = Border(left=Side('thin'),right=Side('thin'),top=Side('thin'),bottom=Side('thin'))
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)

# ═══════════════════════════════════════════════════════
# UTILS
# ═══════════════════════════════════════════════════════

def load_rsu_config():
    if os.path.exists(RSU_CONFIG_PATH):
        try:
            with open(RSU_CONFIG_PATH) as f:
                return json.load(f)
        except: pass
    return {}

def date_diff(d1, d2):
    try:
        for fmt in ['%Y%m%d','%Y-%m-%d','%Y-%m-%d %H:%M:%S']:
            try:
                s1, s2 = d1.replace('-','')[:8], d2.replace('-','')[:8]
                return abs((datetime.strptime(s1,'%Y%m%d')-datetime.strptime(s2,'%Y%m%d')).days)
            except: continue
        return 999
    except: return 999

def fetch_rates(year):
    rates = {}
    for cur in ['USD','HKD']:
        try:
            url = f'https://api.exchangerate-api.com/v4/latest/{cur}'
            req = urllib.request.Request(url, headers={'User-Agent':'Mozilla/5.0'})
            ctx = ssl.create_default_context()
            with urllib.request.urlopen(req, timeout=10, context=ctx) as r:
                rates[cur] = json.loads(r.read().decode())['rates'].get('CNY',0)
        except: pass
    return rates

def detect_file_type(wb):
    """Auto-detect if file is annual CRS report or monthly statement"""
    sheets = set(wb.sheetnames)
    if '证券-交易流水' in sheets:
        return 'annual_crs'
    if any('交易' in s for s in sheets):
        return 'monthly_statement'
    return 'unknown'

# ═══════════════════════════════════════════════════════
# PARSERS — 兼容多种格式
# ═══════════════════════════════════════════════════════

def find_sheet(wb, keywords):
    """Find sheet by keywords (fuzzy match)"""
    for sn in wb.sheetnames:
        if all(kw in sn for kw in keywords):
            return sn
    # Try partial match
    for sn in wb.sheetnames:
        if any(kw in sn for kw in keywords):
            return sn
    return None

def parse_trading_sheet(ws):
    """Parse any trading-like sheet. Auto-detect columns by header name."""
    headers = {}
    for col in range(1, ws.max_column + 1):
        h = str(ws.cell(row=1, column=col).value or '').strip()
        headers[h] = col

    # Column position helpers
    def col(keywords, fallback):
        for kw in keywords:
            for h, c in headers.items():
                if kw in h: return c
        return fallback
    tc = col(['成交时间','时间','日期'], 1)
    catc = col(['品类','类型'], 4)
    codec = col(['代码名称','代码','名称'], 5)
    mktc = col(['交易所/市场','市场'], 6)
    dirc = col(['方向'], 7)
    curc = col(['币种','货币'], 9)
    qtyc = col(['数量/面值','数量'], 10)
    prcc = col(['价格','成交价'], 11)
    amtc = col(['成交金额','金额'], 12)
    feec = col(['总费用','费用'], 13)
    netc = col(['变动金额','净额'], 14)

    records = []
    for row in range(2, ws.max_row + 1):
        time_str = ws.cell(row=row, column=tc).value
        if not time_str: continue

        rec = {
            'time': str(time_str),
            'category': str(ws.cell(row=row, column=catc).value or ''),
            'code': str(ws.cell(row=row, column=codec).value or '').strip(),
            'market': str(ws.cell(row=row, column=mktc).value or '').strip(),
            'direction': str(ws.cell(row=row, column=dirc).value or '').strip(),
            'currency': str(ws.cell(row=row, column=curc).value or '').strip(),
            'qty': float(ws.cell(row=row, column=qtyc).value or 0),
            'price': float(ws.cell(row=row, column=prcc).value or 0),
            'amount': float(ws.cell(row=row, column=amtc).value or 0),
            'fee': abs(float(ws.cell(row=row, column=feec).value or 0)),
            'net_amount': float(ws.cell(row=row, column=netc).value or 0),
            'row': row,
        }
        records.append(rec)
    records.sort(key=lambda r: r['time'])
    return records

def parse_asset_sheet(ws):
    """Parse asset movement sheet"""
    movements = []
    for row in range(2, ws.max_row + 1):
        date = ws.cell(row=row, column=1).value
        if not date: continue
        movements.append({
            'date': str(date),
            'account_name': str(ws.cell(row=row, column=2).value or ''),
            'account_num': str(ws.cell(row=row, column=3).value or ''),
            'category': str(ws.cell(row=row, column=4).value or ''),
            'code': str(ws.cell(row=row, column=5).value or '').strip(),
            'market': str(ws.cell(row=row, column=6).value or '').strip(),
            'direction': str(ws.cell(row=row, column=7).value or ''),
            'type': str(ws.cell(row=row, column=8).value or ''),
            'currency': str(ws.cell(row=row, column=9).value or '').strip(),
            'qty': float(ws.cell(row=row, column=10).value or 0),
            'remark': str(ws.cell(row=row, column=11).value or ''),
        })
    return movements

def parse_fund_flow_sheet(ws):
    """Parse fund flow sheet, classify entries"""
    result = {'dividends':[],'interests':[],'interest_charges':[],'ipo_payments':[],
              'rsu_tax':[],'note_coupons':[],'deposits':[],'other':[]}
    for row in range(2, ws.max_row + 1):
        date = ws.cell(row=row, column=1).value
        if not date: continue
        entry = {
            'date': str(date),
            'account_name': str(ws.cell(row=row, column=2).value or ''),
            'account_num': str(ws.cell(row=row, column=3).value or ''),
            'type': str(ws.cell(row=row, column=4).value or ''),
            'direction': str(ws.cell(row=row, column=5).value or ''),
            'currency': str(ws.cell(row=row, column=6).value or '').strip(),
            'amount': float(ws.cell(row=row, column=7).value or 0),
            'remark': str(ws.cell(row=row, column=8).value or ''),
        }
        t, r, d = entry['type'], entry['remark'].lower(), entry['direction']
        # Dividend detection
        if '分红' in t or 'dividend' in r or ('公司行动' in t and ('dividend' in r or 'i/d' in r.lower() or 's/d' in r.lower())):
            result['dividends'].append(entry)
        elif 'withholding tax' in r or ('公司行动' in t and 'tax' in r):
            result['dividends'].append(entry)
        # Interest
        elif '利息扣除' in t or ('利息' in t and d == 'Out'):
            result['interest_charges'].append(entry)
        elif '利息' in t and d == 'In':
            result['interests'].append(entry)
        # Coupons
        elif 'coupon' in r or '票息' in r:
            result['note_coupons'].append(entry)
        # IPO
        elif 'ipo' in t.lower() or 'ipo' in r:
            result['ipo_payments'].append(entry)
        # RSU tax
        elif any(kw in t for kw in ['激励','长期','收益计划']) or 'rsu' in r:
            result['rsu_tax'].append(entry)
        # Deposits/withdrawals
        elif any(kw in t for kw in ['出入金','资金调拨','账户间','货币兑换']):
            result['deposits'].append(entry)
        else:
            result['other'].append(entry)
    return result

def parse_opening_positions(ws):
    """Parse期初 positions from持仓总览"""
    positions = {}
    for row in range(2, ws.max_row + 1):
        if '期初' not in str(ws.cell(row=row, column=1).value or ''): continue
        code = str(ws.cell(row=row, column=6).value or '').strip()
        market = str(ws.cell(row=row, column=7).value or '').strip()
        currency = str(ws.cell(row=row, column=8).value or '').strip()
        qty = float(ws.cell(row=row, column=9).value or 0)
        price = float(ws.cell(row=row, column=10).value or 0)
        if qty <= 0 or not code: continue
        key = (code, market, currency)
        if key in positions:
            old = positions[key]
            tq = old['qty'] + qty
            old['price'] = (old['price'] * old['qty'] + price * qty) / tq if tq > 0 else 0
            old['qty'] = tq
        else:
            positions[key] = {'qty': qty, 'price': price, 'market': market, 'currency': currency, 'code': code}
    return positions

# ═══════════════════════════════════════════════════════
# ASSET COST FINDER
# ═══════════════════════════════════════════════════════

def find_asset_cost(mv, fund_flows, trading_records, rsu_config=None):
    """Find cost basis for non-trading assets (IPO/RSU/etc.)"""
    mv_type, remark, code = mv['type'], mv['remark'], mv['code']
    mv_date, mv_qty, market = mv['date'], mv['qty'], mv['market']
    result = {'price': 0.0, 'fee': 0.0, 'source': 'unknown'}

    # 0) RSU config lookup
    if rsu_config and code in rsu_config:
        for e in rsu_config[code]:
            if abs(date_diff(e['vest_date'], mv_date)) <= 2 and abs(e['shares'] - mv_qty) < 1:
                result['price'] = e['fmv']
                result['source'] = f'RSU config (FMV={e["fmv"]})'
                return result

    # 1) IPO
    if 'ipo' in mv_type.lower() or 'ipo' in remark.lower():
        for p in fund_flows.get('ipo_payments', []):
            if p['direction'] == 'Out' and abs(date_diff(p['date'], mv_date)) <= 5 and abs(p['amount']) > 100:
                result['price'] = abs(p['amount']) / mv_qty if mv_qty > 0 else 0
                result['source'] = 'IPO payment'
                return result

    # 2) RSU / incentive
    if any(kw in mv_type for kw in ['长期激励','股票收益计划','激励','资产进出']) or 'rsu' in remark.lower():
        # From tax records
        for rsu in fund_flows.get('rsu_tax', []):
            if rsu['direction'] == 'Out' and abs(date_diff(rsu['date'], mv_date)) <= 5:
                tax_amt = abs(rsu['amount'])
                if tax_amt > 0 and mv_qty > 0:
                    for est_rate in [0.10, 0.15, 0.20, 0.22, 0.25, 0.30]:
                        fmv = tax_amt / (est_rate * mv_qty)
                        if 1.0 < fmv < 10000:
                            result['price'] = round(fmv, 4)
                            result['source'] = f'Tax est (rate={est_rate*100:.0f}%)'
                            return result
        # From nearby trades
        if trading_records:
            prices = [rec['price'] for rec in trading_records
                      if rec['code'] == code and rec['market'] == market
                      and abs(date_diff(rec['time'][:8], mv_date)) <= 60]
            if prices:
                result['price'] = round(sum(prices)/len(prices), 4)
                result['source'] = f'Nearby trades avg ({len(prices)})'
                return result
        result['needs_manual'] = True
        return result

    return result

# ═══════════════════════════════════════════════════════
# POSITION TRACKERS
# ═══════════════════════════════════════════════════════

class PositionTrackerWMA:
    """Moving Weighted Average cost tracker"""
    def __init__(self, code, market, currency, category='stock'):
        self.code = code; self.market = market; self.currency = currency
        self.category = category
        self.P = 0.0; self.Q = 0.0; self.total_cost = 0.0; self.realized_pnl = 0.0
        self.total_buy_amt = 0.0; self.total_buy_fee = 0.0
        self.total_sell_amt = 0.0; self.total_sell_fee = 0.0

    def add_opening(self, qty, cost):
        if qty <= 0: return
        self.P = qty; self.Q = cost; self.total_cost = qty * cost

    def buy(self, qty, price, fee=0.0):
        if qty <= 0: return
        buy_cost = qty * price + fee
        self.total_buy_amt += qty * price; self.total_buy_fee += fee
        if self.P > 0:
            self.total_cost += buy_cost; self.P += qty
            self.Q = self.total_cost / self.P
        else:
            self.P = qty; self.total_cost = buy_cost; self.Q = buy_cost / qty

    def sell(self, qty, price, fee=0.0):
        if qty <= 0 or self.P <= 0: return 0.0
        sq = min(qty, self.P)
        realized = (price - self.Q) * sq - fee  # sell fee deducted from P&L
        self.realized_pnl += realized
        self.total_sell_amt += sq * price; self.total_sell_fee += fee
        self.P -= sq; self.total_cost = self.P * self.Q
        if abs(self.P) < 0.000001: self.P = self.Q = self.total_cost = 0.0
        return realized

class PositionTrackerOption:
    """Option position tracker — handles short/long opens and closes"""
    def __init__(self, code, market, currency):
        self.code = code; self.market = market; self.currency = currency
        self.open_positions = []  # [(qty, premium_per_share, direction)]
        self.realized_pnl = 0.0; self.P = 0.0; self.Q = 0.0

    def add_opening(self, qty, cost): pass  # Options don't carry opening positions

    def buy(self, qty, price, fee=0.0):
        """买入开仓(做多) or 买入平仓(平空)"""
        pass  # handled in process_event

    def sell(self, qty, price, fee=0.0):
        """卖出平仓(平多) or 卖出开仓(做空)"""
        pass  # handled in process_event

    def open_short(self, qty, premium, fee=0.0):
        """卖出开仓: receive premium"""
        self.open_positions.append((qty, premium, 'short'))
        self.realized_pnl -= fee  # opening fee is a cost
        self.P += qty  # track open interest

    def close_short(self, qty, premium, fee=0.0):
        """买入平仓: pay premium to close short"""
        realized = 0.0
        remaining = qty
        while remaining > 0 and self.open_positions:
            oq, op, odir = self.open_positions[0]
            if odir != 'short':
                self.open_positions.pop(0)
                continue
            cq = min(remaining, oq)
            realized += (op - premium) * cq  # short: profit = open_premium - close_premium
            remaining -= cq
            if cq >= oq:
                self.open_positions.pop(0)
            else:
                self.open_positions[0] = (oq - cq, op, odir)
        realized -= fee
        self.realized_pnl += realized
        self.P -= qty
        return realized

    def open_long(self, qty, premium, fee=0.0):
        """买入开仓: pay premium"""
        self.open_positions.append((qty, premium, 'long'))
        self.realized_pnl -= fee
        self.P += qty

    def close_long(self, qty, premium, fee=0.0):
        """卖出平仓: receive premium to close long"""
        realized = 0.0
        remaining = qty
        while remaining > 0 and self.open_positions:
            oq, op, odir = self.open_positions[0]
            if odir != 'long':
                self.open_positions.pop(0)
                continue
            cq = min(remaining, oq)
            realized += (premium - op) * cq  # long: profit = close_premium - open_premium
            remaining -= cq
            if cq >= oq:
                self.open_positions.pop(0)
            else:
                self.open_positions[0] = (oq - cq, op, odir)
        realized -= fee
        self.realized_pnl += realized
        self.P -= qty
        return realized

# ═══════════════════════════════════════════════════════
# MAIN CALCULATION ENGINE
# ═══════════════════════════════════════════════════════

def calculate_all(trading_records, asset_movements, fund_flows,
                  opening_positions, method='WMA', rsu_config=None):
    """Main calculation engine — handles stocks, funds, options, futures"""
    trackers = {}
    warnings = []

    def get_tracker(code, market, currency, category='stock'):
        key = (code, market, currency)
        if key not in trackers:
            if 'option' in category.lower() or '期权' in category:
                trackers[key] = PositionTrackerOption(code, market, currency)
            else:
                trackers[key] = PositionTrackerWMA(code, market, currency, category)
        return trackers[key]

    # 1) Load opening positions
    for (code, market, currency), pos in opening_positions.items():
        t = get_tracker(code, market, currency)
        t.add_opening(pos['qty'], pos['price'])

    # 2) Process asset movements (skip account upgrades to avoid double-count)
    asset_entries = []
    for mv in asset_movements:
        if mv['direction'] != 'In': continue
        if any(kw in mv['type'] for kw in ['账户升级','账户迁移','升级']): continue

        cost_info = find_asset_cost(mv, fund_flows, trading_records, rsu_config)
        if cost_info.get('needs_manual'):
            warnings.append({'code': mv['code'], 'market': mv['market'],
                           'qty': mv['qty'], 'type': mv['type'], 'date': mv['date'],
                           'message': f"{mv['code']} {mv['type']}: {mv['qty']:,.0f}股 成本未知"})

        asset_entries.append({
            'time': mv['date'], 'code': mv['code'], 'market': mv['market'],
            'currency': mv['currency'], 'category': mv['category'],
            'qty': mv['qty'], 'cost_per_share': max(cost_info['price'], 0.0001),
            'fee': cost_info.get('fee', 0), 'source': cost_info.get('source', 'unknown'),
        })

    # 3) Merge and sort all events
    all_events = []
    for rec in trading_records:
        all_events.append({'type': 'trade', 'time': rec['time'], 'data': rec})
    for entry in asset_entries:
        all_events.append({'type': 'asset_in', 'time': entry['time'], 'data': entry})
    all_events.sort(key=lambda e: e['time'])

    # 4) Process events
    enriched = []
    for event in all_events:
        if event['type'] == 'asset_in':
            d = event['data']
            t = get_tracker(d['code'], d['market'], d['currency'], d.get('category', 'stock'))
            if d['qty'] > 0:
                t.buy(d['qty'], d['cost_per_share'], d['fee'])
            continue

        rec = event['data']
        category = rec['category']
        t = get_tracker(rec['code'], rec['market'], rec['currency'], category)

        direction = rec['direction']
        qty = abs(rec['qty'])
        price = rec['price']
        fee = rec['fee']

        P_before = t.P if hasattr(t, 'P') else 0
        Q_before = t.Q if hasattr(t, 'Q') else 0
        realized = 0.0

        is_option = '期权' in category or 'option' in category.lower()

        if is_option:
            # ── Option handling ──
            if direction == '卖出开仓':  # Sell to open = short
                t.open_short(qty, price, fee)
            elif direction == '买入平仓':  # Buy to close short
                realized = t.close_short(qty, price, fee)
            elif direction == '买入开仓':  # Buy to open = long
                t.open_long(qty, price, fee)
            elif direction == '卖出平仓':  # Sell to close long
                realized = t.close_long(qty, price, fee)
        else:
            # ── Stock/Fund handling ──
            is_buy = direction in ('买入开仓','买入平仓','申购','买入')
            is_sell = direction in ('卖出平仓','卖出开仓','赎回','卖出')
            if is_buy:
                t.buy(qty, price, fee)
            elif is_sell:
                realized = t.sell(qty, price, fee)

        record = dict(rec)
        record['P_before'] = round(P_before, 4)
        record['Q_before'] = round(Q_before, 4)
        record['P_after'] = round(t.P if hasattr(t, 'P') else 0, 4)
        record['Q_after'] = round(t.Q if hasattr(t, 'Q') else 0, 4)
        record['R'] = round(realized, 4)
        enriched.append(record)

    return enriched, trackers, warnings

# ═══════════════════════════════════════════════════════
# TAX CALCULATION
# ═══════════════════════════════════════════════════════

def calculate_tax(trackers, fund_flows, exchange_rates, year):
    """Calculate tax summary with market breakdown"""
    # Capital gains by market+currency
    market_pnl = defaultdict(lambda: defaultdict(float))
    for (code, market, currency), t in trackers.items():
        pnl = getattr(t, 'realized_pnl', 0)
        if abs(pnl) > 0.005:
            market_pnl[market][currency] += pnl

    # Dividends
    div_by_cur = defaultdict(lambda: {'gross': 0.0, 'tax_withheld': 0.0})
    for div in fund_flows.get('dividends', []):
        r = div['remark'].lower(); cur = div['currency']
        if div['direction'] == 'In' and ('dividend' in r or 'i/d' in r.lower() or 's/d' in r.lower()):
            div_by_cur[cur]['gross'] += abs(div['amount'])
        elif div['direction'] == 'Out' and ('withholding tax' in r or 'tax' in r):
            div_by_cur[cur]['tax_withheld'] += abs(div['amount'])

    # Interest/coupons
    int_by_cur = defaultdict(float)
    for i in fund_flows.get('interests', []):
        if i['direction'] == 'In': int_by_cur[i['currency']] += abs(i['amount'])
    for c in fund_flows.get('note_coupons', []):
        if c['direction'] == 'In': int_by_cur[c['currency']] += abs(c['amount'])

    def to_rmb(amt, cur):
        return amt * exchange_rates.get(cur, 1.0)

    result = {'year': year, 'exchange_rates': exchange_rates,
              'capital_gains': {'by_market': {}, 'total_tax_rmb': 0.0},
              'dividends': {'by_currency': {}, 'total_tax_rmb': 0.0},
              'interests': {'by_currency': {}, 'total_tax_rmb': 0.0}}

    # Capital gains
    for market, curs in sorted(market_pnl.items()):
        label = MARKET_MAP.get(market, f'🌐 {market}')
        me = {'label': label, 'currencies': {}, 'subtotal_tax': 0.0}
        for cur, pnl in sorted(curs.items()):
            pnl_rmb = to_rmb(pnl, cur)
            tax = max(0, pnl_rmb * TAX_RATE)
            me['currencies'][cur] = {'pnl': round(pnl, 2), 'pnl_rmb': round(pnl_rmb, 2),
                                     'rate': exchange_rates.get(cur, 1.0), 'tax_rmb': round(tax, 2)}
            me['subtotal_tax'] += tax
        result['capital_gains']['by_market'][market] = me
        result['capital_gains']['total_tax_rmb'] += me['subtotal_tax']
    result['capital_gains']['total_tax_rmb'] = round(result['capital_gains']['total_tax_rmb'], 2)

    # Dividends
    for cur in sorted(div_by_cur):
        d = div_by_cur[cur]
        gr = to_rmb(d['gross'], cur); wt = to_rmb(d['tax_withheld'], cur)
        should = gr * TAX_RATE; owed = max(0, should - wt)
        result['dividends']['by_currency'][cur] = {
            'gross': round(d['gross'], 2), 'gross_rmb': round(gr, 2),
            'tax_withheld': round(d['tax_withheld'], 2), 'tax_withheld_rmb': round(wt, 2),
            'should_tax_rmb': round(should, 2), 'owed_rmb': round(owed, 2)}
        result['dividends']['total_tax_rmb'] += owed
    result['dividends']['total_tax_rmb'] = round(result['dividends']['total_tax_rmb'], 2)

    # Interests
    for cur in sorted(int_by_cur):
        total = int_by_cur[cur]; tr = to_rmb(total, cur); tax = tr * TAX_RATE
        result['interests']['by_currency'][cur] = {'total': round(total, 2), 'total_rmb': round(tr, 2), 'tax_rmb': round(tax, 2)}
        result['interests']['total_tax_rmb'] += tax
    result['interests']['total_tax_rmb'] = round(result['interests']['total_tax_rmb'], 2)

    result['grand_total_rmb'] = round(
        result['capital_gains']['total_tax_rmb'] + result['dividends']['total_tax_rmb'] + result['interests']['total_tax_rmb'], 2)
    return result

# ═══════════════════════════════════════════════════════
# EXCEL OUTPUT
# ═══════════════════════════════════════════════════════

def apply_style(cell, font=None, fill=None, alignment=None, border=None):
    if font: cell.font = font
    if fill: cell.fill = fill
    if alignment: cell.alignment = alignment
    if border: cell.border = border

def write_output(enriched, trackers, tax_summary, fund_flows, warnings,
                 exchange_rates, output_path, year, method):
    wb = openpyxl.Workbook()

    # Sheet 1: Trading detail with PQR
    ws1 = wb.active
    ws1.title = "📊 交易明细"
    headers = ['成交时间','品类','代码','市场','方向','币种','数量','价格','成交金额','总费用','变动金额',
               'P(前)','Q(前)','P(后)','Q(后)','R(盈亏)']
    for c, h in enumerate(headers, 1):
        apply_style(ws1.cell(row=1, column=c, value=h), HDR_FONT, HDR_FILL, CENTER, THIN_BORDER)
    for i, rec in enumerate(enriched):
        row = i + 2
        vals = [rec['time'][:19], rec['category'], rec['code'], rec['market'], rec['direction'],
                rec['currency'], rec['qty'], rec['price'], rec['amount'], rec['fee'],
                rec['net_amount'], rec['P_before'], rec['Q_before'], rec['P_after'], rec['Q_after'], rec['R']]
        for c, v in enumerate(vals, 1):
            apply_style(ws1.cell(row=row, column=c, value=v), NORM_FONT, None, CENTER, THIN_BORDER)
        r_cell = ws1.cell(row=row, column=16)
        if rec['R'] > 0.005: r_cell.fill = GRN_FILL
        elif rec['R'] < -0.005: r_cell.fill = RED_FILL
    for i, w in enumerate([20,8,10,8,10,8,12,10,14,10,12,10,10,10,10,12], 1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    ws1.freeze_panes = 'A2'
    ws1.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(enriched)+1}"

    # Sheet 2: Holdings summary
    ws2 = wb.create_sheet("📋 持仓汇总")
    ph = ['代码','市场','币种','品类','剩余量','加权成本(Q)','总成本','已实现盈亏','买入额','买入费','卖出额','卖出费']
    for c, h in enumerate(ph, 1):
        apply_style(ws2.cell(row=1, column=c, value=h), HDR_FONT, HDR_FILL, CENTER, THIN_BORDER)
    for i, ((code, market, currency), t) in enumerate(sorted(trackers.items(),
            key=lambda x: abs(getattr(x[1], 'realized_pnl', 0)), reverse=True)):
        row = i + 2
        pnl = getattr(t, 'realized_pnl', 0)
        vals = [code, market, currency, getattr(t, 'category', 'stock'),
                round(getattr(t, 'P', 0), 2), round(getattr(t, 'Q', 0), 4),
                round(getattr(t, 'P', 0) * getattr(t, 'Q', 0), 2), round(pnl, 2),
                round(getattr(t, 'total_buy_amt', 0), 2), round(getattr(t, 'total_buy_fee', 0), 2),
                round(getattr(t, 'total_sell_amt', 0), 2), round(getattr(t, 'total_sell_fee', 0), 2)]
        for c, v in enumerate(vals, 1):
            apply_style(ws2.cell(row=row, column=c, value=v), NORM_FONT, None, CENTER, THIN_BORDER)
        if pnl > 0.005: ws2.cell(row=row, column=8).fill = GRN_FILL
        elif pnl < -0.005: ws2.cell(row=row, column=8).fill = RED_FILL
    for i, w in enumerate([10,8,8,8,12,12,14,14,14,12,14,12], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = 'A2'

    # Sheet 3: Tax report
    ws3 = wb.create_sheet("💰 税务报告")
    row = 1
    ws3.merge_cells('A1:H1')
    apply_style(ws3.cell(row=1, column=1, value=f"境外所得税报告 — {year}年度 ({method})"), TITLE_FONT)
    ws3.row_dimensions[1].height = 35
    row = 3
    apply_style(ws3.cell(row=row, column=1, value=f"汇率: " + ", ".join(f"{c}={r:.4f}" for c,r in exchange_rates.items())), NORM_FONT)
    row += 2

    # Capital gains
    apply_style(ws3.cell(row=row, column=1, value="一、财产转让所得"), Font(name='Arial', size=13, bold=True))
    ws3.merge_cells(f'A{row}:H{row}'); row += 1
    for c, h in enumerate(['市场','币种','净盈亏','汇率','RMB','税额'], 1):
        apply_style(ws3.cell(row=row, column=c, value=h), HDR_FONT, SUB_FILL, CENTER, THIN_BORDER)
    row += 1
    for market, mkt in sorted(tax_summary['capital_gains']['by_market'].items()):
        for cur, d in sorted(mkt['currencies'].items()):
            for c, v in enumerate([mkt['label'], cur, d['pnl'], d['rate'], d['pnl_rmb'], d['tax_rmb']], 1):
                apply_style(ws3.cell(row=row, column=c, value=v), NORM_FONT, None, CENTER, THIN_BORDER)
            if d['tax_rmb'] > 0: ws3.cell(row=row, column=6).font = Font(name='Arial', size=10, bold=True, color='FF0000')
            row += 1
    apply_style(ws3.cell(row=row, column=1, value="小计"), BOLD_FONT, YEL_FILL, CENTER, THIN_BORDER)
    for c in range(2,5): apply_style(ws3.cell(row=row, column=c, value=""), BOLD_FONT, YEL_FILL, CENTER, THIN_BORDER)
    apply_style(ws3.cell(row=row, column=6, value=tax_summary['capital_gains']['total_tax_rmb']), RED_BOLD, YEL_FILL, CENTER, THIN_BORDER)
    row += 2

    # Dividends & interest
    apply_style(ws3.cell(row=row, column=1, value="二、股息利息红利"), Font(name='Arial', size=13, bold=True))
    ws3.merge_cells(f'A{row}:H{row}'); row += 1
    for c, h in enumerate(['类型','币种','收入','RMB','预扣税','预扣RMB','应缴RMB','应补RMB'], 1):
        apply_style(ws3.cell(row=row, column=c, value=h), HDR_FONT, SUB_FILL, CENTER, THIN_BORDER)
    row += 1
    for cur, d in sorted(tax_summary['dividends']['by_currency'].items()):
        for c, v in enumerate(['股息', cur, d['gross'], d['gross_rmb'], d['tax_withheld'], d['tax_withheld_rmb'], d['should_tax_rmb'], d['owed_rmb']], 1):
            apply_style(ws3.cell(row=row, column=c, value=v), NORM_FONT, None, CENTER, THIN_BORDER)
        row += 1
    for cur, d in sorted(tax_summary['interests']['by_currency'].items()):
        for c, v in enumerate(['利息/票息', cur, d['total'], d['total_rmb'], 0, 0, d['tax_rmb'], d['tax_rmb']], 1):
            apply_style(ws3.cell(row=row, column=c, value=v), NORM_FONT, None, CENTER, THIN_BORDER)
        row += 1
    div_int_total = tax_summary['dividends']['total_tax_rmb'] + tax_summary['interests']['total_tax_rmb']
    apply_style(ws3.cell(row=row, column=1, value="小计"), BOLD_FONT, YEL_FILL, CENTER, THIN_BORDER)
    for c in range(2,7): apply_style(ws3.cell(row=row, column=c, value=""), BOLD_FONT, YEL_FILL, CENTER, THIN_BORDER)
    apply_style(ws3.cell(row=row, column=8, value=round(div_int_total, 2)), RED_BOLD, YEL_FILL, CENTER, THIN_BORDER)
    row += 2

    # Total
    ws3.merge_cells(f'A{row}:H{row}')
    apply_style(ws3.cell(row=row, column=1, value=f"三、合计应补税: ¥{tax_summary['grand_total_rmb']:,.2f}"), Font(name='Arial', size=14, bold=True, color='FF0000'))
    ws3.row_dimensions[row].height = 35
    row += 2

    # Reminders
    apply_style(ws3.cell(row=row, column=1, value="⚠️ 重要提醒"), Font(name='Arial', size=11, bold=True, color='FF0000'))
    row += 1
    for r in ["• 申报时间: 次年3月1日-6月30日, 逾期加收日万分之五滞纳金",
              "• 浮盈不征税: 仅实际卖出盈利计税",
              "• 跨年亏损不能结转抵扣",
              "• 境外预扣税可申请抵免",
              f"• 汇率: {year}年12月31日人民币中间价"]:
        apply_style(ws3.cell(row=row, column=1, value=r), NORM_FONT); row += 1

    for i, w in enumerate([30,8,14,10,14,14,14,14], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    # Sheet 4: Dividend detail
    if fund_flows.get('dividends') or fund_flows.get('note_coupons'):
        ws4 = wb.create_sheet("💵 股息利息")
        for c, h in enumerate(['日期','账户','类型','方向','币种','金额','备注'], 1):
            apply_style(ws4.cell(row=1, column=c, value=h), HDR_FONT, HDR_FILL, CENTER, THIN_BORDER)
        r = 2
        for item in fund_flows.get('dividends', []) + fund_flows.get('note_coupons', []):
            for c, k in enumerate(['date','account_name','type','direction','currency','amount','remark'], 1):
                apply_style(ws4.cell(row=r, column=c, value=item[k]), NORM_FONT, None, LEFT, THIN_BORDER)
            r += 1

    # Sheet 5: Warnings
    if warnings:
        ws5 = wb.create_sheet("⚠️ 需确认")
        for c, h in enumerate(['代码','市场','类型','日期','数量','说明'], 1):
            apply_style(ws5.cell(row=1, column=c, value=h), HDR_FONT, HDR_FILL, CENTER, THIN_BORDER)
        for i, w in enumerate(warnings):
            for c, v in enumerate([w['code'], w['market'], w['type'], w['date'], w['qty'], w['message']], 1):
                apply_style(ws5.cell(row=i+2, column=c, value=v), NORM_FONT, YEL_FILL, LEFT, THIN_BORDER)

    wb.save(output_path)
    return output_path

# ═══════════════════════════════════════════════════════
# SINGLE YEAR PROCESSOR
# ═══════════════════════════════════════════════════════

def process_single_year(input_files, prior_year_file=None, output_file=None,
                        year=None, method='WMA', exchange_rates=None):
    """Process single year. input_files can be a single file or list of files (merged)."""
    if isinstance(input_files, str):
        input_files = [input_files]

    if year is None:
        match = re.search(r'(\d{4})', os.path.basename(input_files[0]))
        year = int(match.group(1)) if match else datetime.now().year

    print(f"\n{'='*60}")
    print(f"  📊 {year}年度 | {len(input_files)}个文件 | {method}")
    print(f"{'='*60}")

    # Parse all input files
    all_trading, all_assets, all_flows = [], [], []
    combined_flows = {'dividends':[],'interests':[],'interest_charges':[],'ipo_payments':[],
                      'rsu_tax':[],'note_coupons':[],'deposits':[],'other':[]}

    for f in input_files:
        print(f"  读取: {os.path.basename(f)}")
        wb = openpyxl.load_workbook(f, data_only=True)
        ft = detect_file_type(wb)

        # Find trading sheet
        trade_sn = find_sheet(wb, ['证券','交易流水']) or find_sheet(wb, ['交易'])
        if trade_sn:
            all_trading.extend(parse_trading_sheet(wb[trade_sn]))

        # Find asset movement sheet
        asset_sn = find_sheet(wb, ['证券','资产进出']) or find_sheet(wb, ['资产'])
        if asset_sn:
            all_assets.extend(parse_asset_sheet(wb[asset_sn]))

        # Find fund flow sheet
        flow_sn = find_sheet(wb, ['证券','资金进出']) or find_sheet(wb, ['资金'])
        if flow_sn:
            flows = parse_fund_flow_sheet(wb[flow_sn])
            for k in combined_flows:
                combined_flows[k].extend(flows.get(k, []))

    # Deduplicate (same date+code+direction+qty)
    all_trading = list({(r['time'],r['code'],r['direction'],r['qty']): r for r in all_trading}.values())
    all_trading.sort(key=lambda r: r['time'])
    all_assets.sort(key=lambda d: d['date'])

    print(f"  交易: {len(all_trading)}条 | 资产变动: {len(all_assets)}条 | 股息: {len(combined_flows['dividends'])}条")

    # Load RSU config
    rsu_config = load_rsu_config()

    # Opening positions from prior year
    opening_positions = {}
    if prior_year_file:
        print(f"  读取上年: {os.path.basename(prior_year_file)}")
        wb_prior = openpyxl.load_workbook(prior_year_file, data_only=True)
        prior_trade = parse_trading_sheet(wb_prior[find_sheet(wb_prior, ['证券','交易流水'])])
        prior_asset = parse_asset_sheet(wb_prior[find_sheet(wb_prior, ['证券','资产进出'])])
        prior_flow = parse_fund_flow_sheet(wb_prior[find_sheet(wb_prior, ['证券','资金进出'])])
        prior_open = parse_opening_positions(wb_prior[find_sheet(wb_prior, ['持仓总览'])])

        _, prior_trackers, _ = calculate_all(prior_trade, prior_asset, prior_flow, prior_open, method, rsu_config)
        for key, t in prior_trackers.items():
            if t.P > 0.001:
                code, market, currency = key
                opening_positions[(code, market, currency)] = {
                    'qty': t.P, 'price': t.Q, 'market': market, 'currency': currency, 'code': code}
        print(f"  期初持仓: {len(opening_positions)}个")

    # Exchange rates
    if exchange_rates is None:
        exchange_rates = fetch_rates(year)
        if not exchange_rates:
            exchange_rates = {'USD': 7.0, 'HKD': 0.9}
    for cur, rate in exchange_rates.items():
        print(f"  {cur}/CNY = {rate:.4f}")

    # Calculate
    enriched, trackers, warnings = calculate_all(all_trading, all_assets, combined_flows, opening_positions, method, rsu_config)

    total_R = sum(r['R'] for r in enriched)
    print(f"  R合计: {total_R:,.2f} ({sum(1 for r in enriched if r['R']>0.005)}盈/{sum(1 for r in enriched if r['R']<-0.005)}亏) | {len(trackers)}品种")

    tax = calculate_tax(trackers, combined_flows, exchange_rates, year)

    # Output
    if output_file is None:
        output_file = f"{year}_税务审计底稿_{method}.xlsx"
    write_output(enriched, trackers, tax, combined_flows, warnings, exchange_rates, output_file, year, method)

    print(f"  ✅ {output_file}")
    print(f"  资本利得税: ¥{tax['capital_gains']['total_tax_rmb']:,.2f}")
    print(f"  股息利息税: ¥{tax['dividends']['total_tax_rmb'] + tax['interests']['total_tax_rmb']:,.2f}")
    print(f"  合计应补税: ¥{tax['grand_total_rmb']:,.2f}")

    return {'year': year, 'output': output_file, 'tax': tax, 'total_R': total_R, 'warnings': warnings}

# ═══════════════════════════════════════════════════════
# BATCH PROCESSOR
# ═══════════════════════════════════════════════════════

def process_all_years(directory, method='WMA', exchange_rates=None):
    import glob
    files = sorted(glob.glob(os.path.join(directory, '*年度账单*.xlsx')))
    if not files:
        print("未找到 *年度账单*.xlsx")
        return None
    years = sorted([(int(re.search(r'(\d{4})', os.path.basename(f)).group(1)), f) for f in files])
    if exchange_rates is None and years:
        exchange_rates = fetch_rates(years[-1][0])
    results = []
    for i, (year, fp) in enumerate(years):
        prior = years[i-1][1] if i > 0 else None
        out = os.path.join(directory, 'output', f'{year}_税务审计底稿_{method}.xlsx')
        os.makedirs(os.path.dirname(out), exist_ok=True)
        results.append(process_single_year(fp, prior, out, year, method, exchange_rates))

    # Summary report
    if results:
        _write_summary(results, directory, method, exchange_rates)
    return results

def _write_summary(results, directory, method, exchange_rates):
    out = os.path.join(directory, 'output', f'所有年度汇总_{method}.xlsx')
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "年度汇总"
    apply_style(ws.cell(row=1, column=1, value="境外所得税多年度汇总"), TITLE_FONT)
    ws.merge_cells('A1:I1'); ws.row_dimensions[1].height = 35
    row = 3
    for c, h in enumerate(['年度','资本利得税','股息红利税','利息税','合计','港股盈亏','美股盈亏','交易数','备注'], 1):
        apply_style(ws.cell(row=row, column=c, value=h), HDR_FONT, HDR_FILL, CENTER, THIN_BORDER)
    row += 1
    gtotal = 0
    for r in results:
        t = r['tax']
        hk_pnl = sum(d['pnl'] for m, mkt in t['capital_gains']['by_market'].items() if '港股' in mkt['label'] for d in mkt['currencies'].values())
        us_pnl = sum(d['pnl'] for m, mkt in t['capital_gains']['by_market'].items() if '美股' in mkt['label'] for d in mkt['currencies'].values())
        vals = [r['year'], t['capital_gains']['total_tax_rmb'], t['dividends']['total_tax_rmb'],
                t['interests']['total_tax_rmb'], t['grand_total_rmb'], round(hk_pnl,2), round(us_pnl,2),
                len(r.get('enriched',[])), f"⚠️{len(r['warnings'])}" if r.get('warnings') else '✓']
        for c, v in enumerate(vals, 1):
            apply_style(ws.cell(row=row, column=c, value=v), NORM_FONT, None, CENTER, THIN_BORDER)
        gtotal += t['grand_total_rmb']; row += 1
    apply_style(ws.cell(row=row, column=1, value="合计"), BOLD_FONT, YEL_FILL, CENTER, THIN_BORDER)
    apply_style(ws.cell(row=row, column=5, value=round(gtotal,2)), RED_BOLD, YEL_FILL, CENTER, THIN_BORDER)
    for c in range(2,9): apply_style(ws.cell(row=row, column=c, value=""), BOLD_FONT, YEL_FILL, CENTER, THIN_BORDER)
    for i, w in enumerate([8,16,16,14,16,14,14,12,10], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    wb.save(out)
    print(f"\n📊 汇总: {out}")

# ═══════════════════════════════════════════════════════
# CLI
# ═══════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description=f'富途/券商 CRS 税务计算器 v{VERSION}')
    parser.add_argument('input_files', nargs='*', help='账单文件(支持多个合并)')
    parser.add_argument('--prior', help='上年度文件')
    parser.add_argument('-o','--output', help='输出路径')
    parser.add_argument('--year', type=int, help='年度')
    parser.add_argument('--method', choices=['WMA','FIFO'], default='WMA')
    parser.add_argument('--usd-rate', type=float, help='USD/CNY')
    parser.add_argument('--hkd-rate', type=float, help='HKD/CNY')
    parser.add_argument('--all-years', metavar='DIR', help='批量处理目录')

    args = parser.parse_args()

    rates = {}
    if args.usd_rate: rates['USD'] = args.usd_rate
    if args.hkd_rate: rates['HKD'] = args.hkd_rate
    if not rates: rates = None

    if args.all_years:
        process_all_years(args.all_years, args.method, rates)
        return 0

    if not args.input_files:
        parser.print_help()
        return 1

    process_single_year(args.input_files, args.prior, args.output, args.year, args.method, rates)
    return 0

if __name__ == '__main__':
    sys.exit(main())
