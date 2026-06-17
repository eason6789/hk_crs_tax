#!/usr/bin/env python3
"""
╔══════════════════════════════════════════════════════════════════╗
║       富途 CRS 境外所得税计算器 v2.0                              ║
║       Futu CRS Offshore Income Tax Calculator                    ║
║                                                                  ║
║  功能：解析富途年度账单 Excel，按移动加权平均法/FIFO 计算          ║
║        股票/基金/期权转让所得，股息红利，利息收入，                ║
║        输出分市场(港股/美股/其他)的应补税额。                      ║
║                                                                  ║
║  用法：                                                           ║
║    python3 futu_tax_calculator.py <年度账单.xlsx> [上年账单.xlsx] ║
║    python3 futu_tax_calculator.py --all-years <目录>              ║
║                                                                  ║
║  法律依据：                                                       ║
║    - 《个人所得税法》(2018修正)                                   ║
║    - 财税〔2016〕36号 — 金融商品转让成本核算                      ║
║    - 财税2020年第3号 — 境外所得申报                               ║
╚══════════════════════════════════════════════════════════════════╝
"""

import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                               numbers, NamedStyle)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
import sys, os, json, re, argparse
from datetime import datetime, date
from collections import defaultdict
from copy import copy
import urllib.request
import ssl

# ═══════════════════════════════════════════════════════════════════
# 0. 常量与配置
# ═══════════════════════════════════════════════════════════════════

VERSION = "2.0.0"

# 市场分类 (用于分市场统计)
MARKET_CLASSIFICATION = {
    'US': '🇺🇸 美股',
    'SEHK': '🇭🇰 港股',
    'FD': '🌐 基金',
    '-': '🌐 其他',
    'FTSN': '🌐 结构化票据',
}

TAX_RATE = 0.20  # 财产转让所得/股息红利统一税率

# Excel 样式
COLORS = {
    'primary': '1F4E79',
    'secondary': '2E75B6',
    'accent': '4472C4',
    'profit_green': 'C6EFCE',
    'loss_red': 'FFC7CE',
    'warning_yellow': 'FFE699',
    'light_gray': 'D6DCE4',
    'white': 'FFFFFF',
    'dark': '333333',
}

HEADER_FILL = PatternFill(start_color=COLORS['primary'], end_color=COLORS['primary'], fill_type='solid')
SUBHEADER_FILL = PatternFill(start_color=COLORS['secondary'], end_color=COLORS['secondary'], fill_type='solid')
GREEN_FILL = PatternFill(start_color=COLORS['profit_green'], end_color=COLORS['profit_green'], fill_type='solid')
RED_FILL = PatternFill(start_color=COLORS['loss_red'], end_color=COLORS['loss_red'], fill_type='solid')
YELLOW_FILL = PatternFill(start_color=COLORS['warning_yellow'], end_color=COLORS['warning_yellow'], fill_type='solid')
HEADER_FONT = Font(name='微软雅黑', size=11, bold=True, color=COLORS['white'])
TITLE_FONT = Font(name='微软雅黑', size=16, bold=True, color=COLORS['primary'])
SUBTITLE_FONT = Font(name='微软雅黑', size=12, bold=True, color=COLORS['dark'])
NORMAL_FONT = Font(name='微软雅黑', size=10)
BOLD_FONT = Font(name='微软雅黑', size=10, bold=True)
RED_BOLD = Font(name='微软雅黑', size=12, bold=True, color='FF0000')
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_ALIGN = Alignment(horizontal='left', vertical='center', wrap_text=True)


# ═══════════════════════════════════════════════════════════════════
# 1. 汇率获取
# ═══════════════════════════════════════════════════════════════════

def fetch_exchange_rates(year):
    """获取纳税年度最后一日的人民币汇率中间价"""
    rates = {}
    date_str = f"{year}-12-31"

    # Try multiple free APIs
    apis = [
        # exchangerate-api (free, reliable)
        lambda curr: f"https://api.exchangerate-api.com/v4/latest/{curr}",
        # frankfurter (free, ECB rates)
        lambda curr: f"https://api.frankfurter.app/{date_str}?from={curr}&to=CNY",
    ]

    for currency in ['USD', 'HKD']:
        for api_func in apis:
            try:
                url = api_func(currency)
                ctx = ssl.create_default_context()
                req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
                with urllib.request.urlopen(req, timeout=10, context=ctx) as resp:
                    data = json.loads(resp.read().decode())
                    if 'rates' in data:
                        rates[currency] = data['rates'].get('CNY', 0)
                    else:
                        rates[currency] = data.get('rates', {}).get('CNY', 0)
                if rates.get(currency):
                    break
            except Exception:
                continue

    return rates


# ═══════════════════════════════════════════════════════════════════
# 2. Excel 解析
# ═══════════════════════════════════════════════════════════════════

def parse_trading_records(ws):
    """解析证券-交易流水"""
    records = []
    for row in range(2, ws.max_row + 1):
        time_str = ws.cell(row=row, column=1).value
        if not time_str:
            continue
        records.append({
            'time': str(time_str),
            'account_name': str(ws.cell(row=row, column=2).value or ''),
            'account_num': str(ws.cell(row=row, column=3).value or ''),
            'category': str(ws.cell(row=row, column=4).value or ''),
            'code': str(ws.cell(row=row, column=5).value or '').strip(),
            'market': str(ws.cell(row=row, column=6).value or '').strip(),
            'direction': str(ws.cell(row=row, column=7).value or '').strip(),
            'settle_date': str(ws.cell(row=row, column=8).value or ''),
            'currency': str(ws.cell(row=row, column=9).value or '').strip(),
            'qty': float(ws.cell(row=row, column=10).value or 0),
            'price': float(ws.cell(row=row, column=11).value or 0),
            'amount': float(ws.cell(row=row, column=12).value or 0),
            'fee': abs(float(ws.cell(row=row, column=13).value or 0)),
            'net_amount': float(ws.cell(row=row, column=14).value or 0),
            'row': row,
        })
    records.sort(key=lambda r: r['time'])
    return records


def parse_asset_movements(ws):
    """解析证券-资产进出"""
    movements = []
    for row in range(2, ws.max_row + 1):
        mv_date = ws.cell(row=row, column=1).value
        if not mv_date:
            continue
        movements.append({
            'date': str(mv_date),
            'account_name': str(ws.cell(row=row, column=2).value or ''),
            'account_num': str(ws.cell(row=row, column=3).value or ''),
            'category': str(ws.cell(row=row, column=4).value or ''),
            'code': str(ws.cell(row=row, column=5).value or '').strip(),
            'market': str(ws.cell(row=row, column=6).value or '').strip(),
            'direction': str(ws.cell(row=row, column=7).value or ''),
            'type': str(ws.cell(row=row, column=8).value or ''),
            'currency': str(ws.cell(row=row, column=9).value or '').strip(),
            'qty': float(ws.cell(row=row, column=10).value or 0),
            'remark': str(ws.cell(row=row, column=11).value or ''),
        })
    return movements


def parse_fund_flows(ws):
    """解析证券-资金进出，分类识别"""
    result = {
        'dividends': [],
        'interests': [],
        'interest_charges': [],  # 融资利息支出
        'ipo_payments': [],
        'rsu_tax': [],
        'note_coupons': [],
        'deposits_withdrawals': [],
        'other': [],
    }

    for row in range(2, ws.max_row + 1):
        f_date = ws.cell(row=row, column=1).value
        if not f_date:
            continue

        entry = {
            'date': str(f_date),
            'account_name': str(ws.cell(row=row, column=2).value or ''),
            'account_num': str(ws.cell(row=row, column=3).value or ''),
            'type': str(ws.cell(row=row, column=4).value or ''),
            'direction': str(ws.cell(row=row, column=5).value or ''),
            'currency': str(ws.cell(row=row, column=6).value or '').strip(),
            'amount': float(ws.cell(row=row, column=7).value or 0),
            'remark': str(ws.cell(row=row, column=8).value or ''),
        }

        t = entry['type']
        r = entry['remark'].lower()
        d = entry['direction']

        # 股息红利识别
        if ('分红' in t or 'dividend' in r or
            ('公司行动' in t and ('dividend' in r or 'i/d' in r.lower() or 's/d' in r.lower()))):
            result['dividends'].append(entry)
        # 股息预扣税
        elif ('withholding tax' in r or
              ('公司行动' in t and 'tax' in r)):
            result['dividends'].append(entry)
        # 融资利息支出
        elif ('利息扣除' in t or '利息' in t):
            if d == 'Out':
                result['interest_charges'].append(entry)
            else:
                result['interests'].append(entry)
        # 票息
        elif ('coupon' in r or '票息' in r):
            result['note_coupons'].append(entry)
        # IPO
        elif ('ipo' in t.lower() or 'ipo' in r):
            result['ipo_payments'].append(entry)
        # RSU 税款
        elif ('激励' in t or 'rsu' in r or '长期' in t or '收益计划' in t):
            result['rsu_tax'].append(entry)
        # 出入金
        elif ('出入金' in t or '资金调拨' in t or '账户间' in t or '货币兑换' in t):
            result['deposits_withdrawals'].append(entry)
        else:
            result['other'].append(entry)

    return result


def parse_opening_positions(ws, period='期初'):
    """解析持仓总览的期初/期末持仓"""
    positions = {}
    for row in range(2, ws.max_row + 1):
        period_type = str(ws.cell(row=row, column=1).value or '')
        if period not in period_type:
            continue
        code = str(ws.cell(row=row, column=6).value or '').strip()
        market = str(ws.cell(row=row, column=7).value or '').strip()
        currency = str(ws.cell(row=row, column=8).value or '').strip()
        qty = float(ws.cell(row=row, column=9).value or 0)
        price = float(ws.cell(row=row, column=10).value or 0)

        if qty <= 0 or not code:
            continue

        key = (code, market, currency)
        if key in positions:
            old = positions[key]
            total_qty = old['qty'] + qty
            old['price'] = (old['price'] * old['qty'] + price * qty) / total_qty if total_qty > 0 else 0
            old['qty'] = total_qty
        else:
            positions[key] = {'qty': qty, 'price': price, 'market': market,
                              'currency': currency, 'code': code}
    return positions


# ═══════════════════════════════════════════════════════════════════
# 3. 成本计算方法
# ═══════════════════════════════════════════════════════════════════

class PositionTrackerWMA:
    """移动加权平均法 (Moving Weighted Average)"""

    def __init__(self, code, market, currency):
        self.code = code
        self.market = market
        self.currency = currency
        self.P = 0.0
        self.Q = 0.0
        self.total_cost = 0.0
        self.realized_pnl = 0.0
        self.total_buy_amount = 0.0
        self.total_buy_fee = 0.0
        self.total_sell_amount = 0.0
        self.total_sell_fee = 0.0
        self.transactions = 0

    def add_opening(self, qty, cost_per_share):
        if qty <= 0:
            return
        self.P = qty
        self.Q = cost_per_share
        self.total_cost = qty * cost_per_share

    def buy(self, qty, price, fee=0.0):
        if qty <= 0:
            return
        buy_cost = qty * price + fee  # 买入手续费计入成本
        self.total_buy_amount += qty * price
        self.total_buy_fee += fee
        self.transactions += 1

        if self.P > 0:
            self.total_cost += buy_cost
            self.P += qty
            self.Q = self.total_cost / self.P
        else:
            self.P = qty
            self.total_cost = buy_cost
            self.Q = buy_cost / qty

    def sell(self, qty, price, fee=0.0):
        if qty <= 0 or self.P <= 0:
            return 0.0
        sell_qty = min(qty, self.P)
        gross = (price - self.Q) * sell_qty
        realized = gross - fee
        self.realized_pnl += realized
        self.total_sell_amount += sell_qty * price
        self.total_sell_fee += fee
        self.transactions += 1
        self.P -= sell_qty
        self.total_cost = self.P * self.Q
        if abs(self.P) < 0.000001:
            self.P = 0.0
            self.Q = 0.0
            self.total_cost = 0.0
        return realized


class PositionTrackerFIFO:
    """先进先出法 (FIFO)"""

    def __init__(self, code, market, currency):
        self.code = code
        self.market = market
        self.currency = currency
        self.lots = []  # [(qty, price, fee)]
        self.realized_pnl = 0.0
        self.total_buy_amount = 0.0
        self.total_buy_fee = 0.0
        self.total_sell_amount = 0.0
        self.total_sell_fee = 0.0

    @property
    def P(self):
        return sum(lot[0] for lot in self.lots)

    @property
    def Q(self):
        total_qty = self.P
        if total_qty <= 0:
            return 0.0
        total_cost = sum(lot[0] * lot[1] + lot[2] for lot in self.lots)
        return total_cost / total_qty

    def add_opening(self, qty, cost_per_share):
        if qty <= 0:
            return
        self.lots.append((qty, cost_per_share, 0.0))

    def buy(self, qty, price, fee=0.0):
        if qty <= 0:
            return
        self.total_buy_amount += qty * price
        self.total_buy_fee += fee
        self.lots.append((qty, price, fee))

    def sell(self, qty, price, fee=0.0):
        if qty <= 0 or not self.lots:
            return 0.0
        remaining = qty
        realized = 0.0
        while remaining > 0 and self.lots:
            lot_qty, lot_price, lot_fee = self.lots[0]
            use_qty = min(remaining, lot_qty)
            lot_cost_per_share = lot_price + (lot_fee / lot_qty if lot_qty > 0 else 0)
            gross = (price - lot_cost_per_share) * use_qty
            realized += gross
            remaining -= use_qty
            if use_qty >= lot_qty:
                self.lots.pop(0)
            else:
                self.lots[0] = (lot_qty - use_qty, lot_price, lot_fee * (1 - use_qty / lot_qty))

        realized -= fee
        self.realized_pnl += realized
        self.total_sell_amount += qty * price
        self.total_sell_fee += fee
        return realized


# ═══════════════════════════════════════════════════════════════════
# 4. 资产成本查找
# ═══════════════════════════════════════════════════════════════════

def date_diff(d1, d2):
    """计算两个日期字符串的天数差"""
    try:
        for fmt in ['%Y%m%d', '%Y-%m-%d', '%Y-%m-%d %H:%M:%S']:
            try:
                s1 = d1.replace('-', '')[:8]
                s2 = d2.replace('-', '')[:8]
                dt1 = datetime.strptime(s1, '%Y%m%d')
                dt2 = datetime.strptime(s2, '%Y%m%d')
                return abs((dt1 - dt2).days)
            except:
                continue
        return 999
    except:
        return 999


def find_asset_cost(mv, fund_flows, trading_records):
    """
    查找非交易资产(IPO/RSU等)的成本基准
    优先级: IPO扣款 > RSU税款反推 > 附近交易价格 > 0(需手动)
    """
    mv_type = mv['type']
    remark = mv['remark']
    code = mv['code']
    mv_date = mv['date']
    mv_qty = mv['qty']
    market = mv['market']

    result = {'price': 0.0, 'fee': 0.0, 'source': 'unknown'}

    # ── IPO 中签 ──
    if 'ipo' in mv_type.lower() or 'ipo' in remark.lower():
        for payment in fund_flows.get('ipo_payments', []):
            if (payment['direction'] == 'Out' and
                abs(date_diff(payment['date'], mv_date)) <= 5 and
                abs(payment['amount']) > 100):
                result['price'] = abs(payment['amount']) / mv_qty if mv_qty > 0 else 0
                result['source'] = 'IPO扣款记录'
                return result

    # ── RSU/长期激励 ──
    if any(kw in mv_type for kw in ['长期激励', '股票收益计划', '激励']) or 'rsu' in remark.lower():
        # 1) 从税款记录估算
        for rsu in fund_flows.get('rsu_tax', []):
            if rsu['direction'] == 'Out' and abs(date_diff(rsu['date'], mv_date)) <= 5:
                tax_amt = abs(rsu['amount'])
                if tax_amt > 0 and mv_qty > 0:
                    # RSU预扣税款约为 FMV × 20-45%
                    for est_rate in [0.22, 0.30, 0.35, 0.45]:
                        est_fmv = tax_amt / (est_rate * mv_qty)
                        if 1.0 < est_fmv < 10000:
                            result['price'] = round(est_fmv, 4)
                            result['source'] = f'RSU税款估算(税率{est_rate*100:.0f}%)'
                            return result

        # 2) 从附近交易价格估算
        if trading_records:
            prices = []
            for rec in trading_records:
                if rec['code'] == code and rec['market'] == market:
                    try:
                        if abs(date_diff(rec['time'][:8], mv_date)) <= 60:
                            prices.append(rec['price'])
                    except:
                        pass
            if prices:
                avg_px = sum(prices) / len(prices)
                result['price'] = round(avg_px, 4)
                result['source'] = f'附近交易均价({len(prices)}笔)'
                return result

        # 3) 无法自动确定
        result['price'] = 0.0
        result['source'] = '需手动输入'
        result['needs_manual'] = True
        return result

    return result


# ═══════════════════════════════════════════════════════════════════
# 5. 主计算引擎
# ═══════════════════════════════════════════════════════════════════

def calculate_all(trading_records, asset_movements, fund_flows,
                  opening_positions, method='WMA'):
    """
    主计算函数
    method: 'WMA' (移动加权平均) 或 'FIFO' (先进先出)
    """
    trackers = {}
    warnings = []

    def get_tracker(code, market, currency):
        key = (code, market, currency)
        if key not in trackers:
            if method == 'FIFO':
                trackers[key] = PositionTrackerFIFO(code, market, currency)
            else:
                trackers[key] = PositionTrackerWMA(code, market, currency)
        return trackers[key]

    # 1) 加载期初持仓
    for (code, market, currency), pos in opening_positions.items():
        t = get_tracker(code, market, currency)
        t.add_opening(pos['qty'], pos['price'])

    # 2) 处理资产进出（排除账户升级以避免重复）
    asset_entries = []
    for mv in asset_movements:
        if mv['direction'] != 'In':
            continue
        # 账户升级/迁移 = 旧持仓转移，已含在期初中
        if any(kw in mv['type'] for kw in ['账户升级', '账户迁移', '升级']):
            continue

        cost_info = find_asset_cost(mv, fund_flows, trading_records)
        if cost_info.get('needs_manual'):
            warnings.append({
                'code': mv['code'],
                'market': mv['market'],
                'qty': mv['qty'],
                'type': mv['type'],
                'date': mv['date'],
                'message': f"{mv['code']} {mv['type']}: {mv['qty']:,.0f}股成本未知，暂按0计",
            })

        asset_entries.append({
            'time': mv['date'],
            'code': mv['code'],
            'market': mv['market'],
            'currency': mv['currency'],
            'qty': mv['qty'],
            'cost_per_share': max(cost_info['price'], 0.0001),
            'fee': cost_info.get('fee', 0),
            'source': cost_info.get('source', 'unknown'),
            'needs_manual': cost_info.get('needs_manual', False),
        })

    # 3) 合并、排序所有事件
    all_events = []
    for rec in trading_records:
        all_events.append({'type': 'trade', 'time': rec['time'], 'data': rec})

    for entry in asset_entries:
        all_events.append({'type': 'asset_in', 'time': entry['time'], 'data': entry})

    all_events.sort(key=lambda e: e['time'])

    # 4) 逐笔处理
    enriched = []
    for event in all_events:
        if event['type'] == 'asset_in':
            d = event['data']
            t = get_tracker(d['code'], d['market'], d['currency'])
            if d['qty'] > 0:
                t.buy(d['qty'], d['cost_per_share'], d['fee'])
            continue

        rec = event['data']
        t = get_tracker(rec['code'], rec['market'], rec['currency'])

        direction = rec['direction']
        qty = abs(rec['qty'])
        price = rec['price']
        fee = rec['fee']

        P_before = t.P
        Q_before = t.Q
        realized = 0.0

        is_buy = direction in ('买入开仓', '买入平仓', '申购', '买入')
        is_sell = direction in ('卖出平仓', '卖出开仓', '赎回', '卖出')

        if is_buy:
            t.buy(qty, price, fee)
        elif is_sell:
            realized = t.sell(qty, price, fee)

        record = dict(rec)
        record['P_before'] = round(P_before, 4)
        record['Q_before'] = round(Q_before, 4)
        record['P_after'] = round(t.P, 4)
        record['Q_after'] = round(t.Q, 4)
        record['R'] = round(realized, 4)
        enriched.append(record)

    return enriched, trackers, warnings


# ═══════════════════════════════════════════════════════════════════
# 6. 税务计算与汇总
# ═══════════════════════════════════════════════════════════════════

def classify_market(market):
    """将市场代码映射到分类"""
    for key, label in MARKET_CLASSIFICATION.items():
        if key == market:
            return label
    return '🌐 其他'


def calculate_tax(trackers, fund_flows, exchange_rates, year):
    """计算税务汇总，按市场分开"""

    # ── 6a. 资本利得 — 按市场+币种 ──
    market_pnl = defaultdict(lambda: defaultdict(float))

    for (code, market, currency), t in trackers.items():
        pnl = t.realized_pnl if hasattr(t, 'realized_pnl') else getattr(t, 'realized_pnl', 0)
        if abs(pnl) > 0.005:
            market_pnl[market][currency] += pnl

    # ── 6b. 股息 ──
    div_by_currency = defaultdict(lambda: {'gross': 0.0, 'tax_withheld': 0.0})
    for div in fund_flows.get('dividends', []):
        r = div['remark'].lower()
        cur = div['currency']
        if div['direction'] == 'In' and ('dividend' in r or 'i/d' in r.lower() or 's/d' in r.lower()):
            div_by_currency[cur]['gross'] += abs(div['amount'])
        elif div['direction'] == 'Out' and ('withholding tax' in r or 'tax' in r):
            div_by_currency[cur]['tax_withheld'] += abs(div['amount'])

    # ── 6c. 利息/票息 ──
    interest_by_currency = defaultdict(float)
    for interest in fund_flows.get('interests', []):
        if interest['direction'] == 'In':
            interest_by_currency[interest['currency']] += abs(interest['amount'])
    for coupon in fund_flows.get('note_coupons', []):
        if coupon['direction'] == 'In':
            interest_by_currency[coupon['currency']] += abs(coupon['amount'])

    # ── 6d. 折算RMB ──
    def to_rmb(amount, currency):
        rate = exchange_rates.get(currency, 1.0)
        return amount * rate

    result = {
        'year': year,
        'exchange_rates': exchange_rates,
        'capital_gains': {'by_market': {}, 'total_tax_rmb': 0.0},
        'dividends': {'by_currency': {}, 'total_tax_rmb': 0.0},
        'interests': {'by_currency': {}, 'total_tax_rmb': 0.0},
    }

    # 资本利得
    for market, currencies in sorted(market_pnl.items()):
        market_label = classify_market(market)
        market_entry = {'label': market_label, 'currencies': {}, 'subtotal_tax': 0.0}
        for currency, pnl in sorted(currencies.items()):
            pnl_rmb = to_rmb(pnl, currency)
            tax = max(0, pnl_rmb * TAX_RATE)
            market_entry['currencies'][currency] = {
                'pnl': round(pnl, 2),
                'pnl_rmb': round(pnl_rmb, 2),
                'rate': exchange_rates.get(currency, 1.0),
                'tax_rmb': round(tax, 2),
            }
            market_entry['subtotal_tax'] += tax
        result['capital_gains']['by_market'][market] = market_entry
        result['capital_gains']['total_tax_rmb'] += market_entry['subtotal_tax']

    result['capital_gains']['total_tax_rmb'] = round(result['capital_gains']['total_tax_rmb'], 2)

    # 股息
    for currency in sorted(div_by_currency):
        d = div_by_currency[currency]
        gross_rmb = to_rmb(d['gross'], currency)
        withheld_rmb = to_rmb(d['tax_withheld'], currency)
        should_tax = gross_rmb * TAX_RATE
        owed = max(0, should_tax - withheld_rmb)
        result['dividends']['by_currency'][currency] = {
            'gross': round(d['gross'], 2),
            'gross_rmb': round(gross_rmb, 2),
            'tax_withheld': round(d['tax_withheld'], 2),
            'tax_withheld_rmb': round(withheld_rmb, 2),
            'should_tax_rmb': round(should_tax, 2),
            'owed_rmb': round(owed, 2),
        }
        result['dividends']['total_tax_rmb'] += owed

    result['dividends']['total_tax_rmb'] = round(result['dividends']['total_tax_rmb'], 2)

    # 利息
    for currency in sorted(interest_by_currency):
        total = interest_by_currency[currency]
        total_rmb = to_rmb(total, currency)
        tax = total_rmb * TAX_RATE
        result['interests']['by_currency'][currency] = {
            'total': round(total, 2),
            'total_rmb': round(total_rmb, 2),
            'tax_rmb': round(tax, 2),
        }
        result['interests']['total_tax_rmb'] += tax

    result['interests']['total_tax_rmb'] = round(result['interests']['total_tax_rmb'], 2)

    # 总计
    result['grand_total_rmb'] = round(
        result['capital_gains']['total_tax_rmb'] +
        result['dividends']['total_tax_rmb'] +
        result['interests']['total_tax_rmb'], 2
    )

    return result


# ═══════════════════════════════════════════════════════════════════
# 7. Excel 输出生成
# ═══════════════════════════════════════════════════════════════════

def apply_cell_style(cell, font=None, fill=None, alignment=None, border=None, number_format=None):
    if font: cell.font = font
    if fill: cell.fill = fill
    if alignment: cell.alignment = alignment
    if border: cell.border = border
    if number_format: cell.number_format = number_format


def write_output(enriched, trackers, tax_summary, fund_flows, warnings,
                 exchange_rates, output_path, year, method):
    """生成专业格式的 Excel 审计底稿"""
    wb = openpyxl.Workbook()

    # ── Sheet 1: 交易明细(含PQR) ──
    ws1 = wb.active
    ws1.title = "📊 交易明细"

    headers = ['成交时间', '品类', '代码', '市场', '方向', '币种', '数量',
               '价格', '成交金额', '总费用', '变动金额',
               'P(前)', 'Q(前)', 'P(后)', 'Q(后)', 'R(盈亏)']

    for c, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=c, value=h)
        apply_cell_style(cell, HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)

    for i, rec in enumerate(enriched):
        row = i + 2
        vals = [
            rec['time'][:19], rec['category'], rec['code'],
            rec['market'], rec['direction'], rec['currency'],
            rec['qty'], rec['price'], rec['amount'], rec['fee'],
            rec['net_amount'], rec['P_before'], rec['Q_before'],
            rec['P_after'], rec['Q_after'], rec['R'],
        ]
        for c, v in enumerate(vals, 1):
            cell = ws1.cell(row=row, column=c, value=v)
            apply_cell_style(cell, NORMAL_FONT, None, CENTER_ALIGN, THIN_BORDER)

        # R列着色
        r_cell = ws1.cell(row=row, column=16)
        if rec['R'] > 0.005:
            r_cell.fill = GREEN_FILL
        elif rec['R'] < -0.005:
            r_cell.fill = RED_FILL

    col_widths_1 = [20, 8, 10, 8, 10, 8, 12, 10, 14, 10, 12, 10, 10, 10, 10, 12]
    for i, w in enumerate(col_widths_1, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    ws1.freeze_panes = 'A2'

    # 添加自动筛选
    ws1.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(enriched)+1}"

    # ── Sheet 2: 持仓汇总 ──
    ws2 = wb.create_sheet("📋 持仓汇总")
    pos_headers = ['代码', '市场', '币种', '剩余股数', '加权成本(Q)', '总成本',
                   '已实现盈亏', '买入总额', '买入手续费', '卖出总额', '卖出手续费']

    for c, h in enumerate(pos_headers, 1):
        apply_cell_style(ws2.cell(row=1, column=c, value=h),
                         HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)

    sorted_trackers = sorted(trackers.items(),
                             key=lambda x: abs(getattr(x[1], 'realized_pnl', 0)), reverse=True)

    for i, ((code, market, currency), t) in enumerate(sorted_trackers):
        row = i + 2
        vals = [
            code, market, currency,
            round(t.P, 2), round(t.Q, 4), round(t.P * t.Q, 2),
            round(getattr(t, 'realized_pnl', 0), 2),
            round(getattr(t, 'total_buy_amount', 0), 2),
            round(getattr(t, 'total_buy_fee', 0), 2),
            round(getattr(t, 'total_sell_amount', 0), 2),
            round(getattr(t, 'total_sell_fee', 0), 2),
        ]
        for c, v in enumerate(vals, 1):
            apply_cell_style(ws2.cell(row=row, column=c, value=v),
                             NORMAL_FONT, None, CENTER_ALIGN, THIN_BORDER)

        r_cell = ws2.cell(row=row, column=7)  # realized PnL
        if getattr(t, 'realized_pnl', 0) > 0.005:
            r_cell.fill = GREEN_FILL
        elif getattr(t, 'realized_pnl', 0) < -0.005:
            r_cell.fill = RED_FILL

    col_widths_2 = [10, 8, 8, 12, 12, 14, 14, 14, 12, 14, 12]
    for i, w in enumerate(col_widths_2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = 'A2'

    # ── Sheet 3: 税务报告 ──
    ws3 = wb.create_sheet("💰 税务报告")
    row = 1

    # 标题
    ws3.merge_cells('A1:H1')
    c = ws3.cell(row=1, column=1, value=f"境外所得税计算报告 — {year}年度")
    apply_cell_style(c, TITLE_FONT, None, Alignment(horizontal='center', vertical='center'))
    ws3.row_dimensions[1].height = 35

    row = 3
    ws3.cell(row=row, column=1, value=f"计算方法: {'移动加权平均法(WMA)' if method == 'WMA' else '先进先出法(FIFO)'}").font = SUBTITLE_FONT
    row += 1
    ws3.cell(row=row, column=1, value=f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = NORMAL_FONT
    row += 2

    # 汇率
    ws3.cell(row=row, column=1, value="使用汇率 (纳税年度末日中间价):").font = BOLD_FONT
    row += 1
    for cur, rate in exchange_rates.items():
        ws3.cell(row=row, column=1, value=f"  {cur}/CNY = {rate:.4f}").font = NORMAL_FONT
        row += 1
    row += 1

    # ── 一、资本利得 ──
    ws3.merge_cells(f'A{row}:H{row}')
    ws3.cell(row=row, column=1, value="一、财产转让所得（资本利得）").font = Font(name='微软雅黑', size=13, bold=True)
    ws3.row_dimensions[row].height = 28
    row += 1

    cg_headers = ['市场', '币种', '年度净盈亏', '汇率', '折算RMB', '应纳税额(RMB)']
    for c, h in enumerate(cg_headers, 1):
        apply_cell_style(ws3.cell(row=row, column=c, value=h),
                         HEADER_FONT, SUBHEADER_FILL, CENTER_ALIGN, THIN_BORDER)
    row += 1

    cg = tax_summary['capital_gains']
    for market, mkt_data in sorted(cg['by_market'].items()):
        for cur, cur_data in sorted(mkt_data['currencies'].items()):
            vals = [mkt_data['label'], cur, cur_data['pnl'], cur_data['rate'],
                    cur_data['pnl_rmb'], cur_data['tax_rmb']]
            for c, v in enumerate(vals, 1):
                apply_cell_style(ws3.cell(row=row, column=c, value=v),
                                 NORMAL_FONT, None, CENTER_ALIGN, THIN_BORDER)
            if cur_data['tax_rmb'] > 0:
                ws3.cell(row=row, column=6).font = Font(name='微软雅黑', size=10, bold=True, color='FF0000')
            row += 1

    # 小计
    apply_cell_style(ws3.cell(row=row, column=1, value="资本利得税小计"), BOLD_FONT, YELLOW_FILL, CENTER_ALIGN, THIN_BORDER)
    for c in range(2, 5):
        apply_cell_style(ws3.cell(row=row, column=c, value=""), BOLD_FONT, YELLOW_FILL, CENTER_ALIGN, THIN_BORDER)
    apply_cell_style(ws3.cell(row=row, column=5, value=""), BOLD_FONT, YELLOW_FILL, CENTER_ALIGN, THIN_BORDER)
    apply_cell_style(ws3.cell(row=row, column=6, value=cg['total_tax_rmb']), RED_BOLD, YELLOW_FILL, CENTER_ALIGN, THIN_BORDER)
    row += 2

    # ── 二、股息红利 ──
    ws3.merge_cells(f'A{row}:H{row}')
    ws3.cell(row=row, column=1, value="二、利息、股息、红利所得").font = Font(name='微软雅黑', size=13, bold=True)
    ws3.row_dimensions[row].height = 28
    row += 1

    div_headers = ['类型', '币种', '收入总额', '收入RMB', '预扣税', '预扣税RMB', '应缴RMB', '应补RMB']
    for c, h in enumerate(div_headers, 1):
        apply_cell_style(ws3.cell(row=row, column=c, value=h),
                         HEADER_FONT, SUBHEADER_FILL, CENTER_ALIGN, THIN_BORDER)
    row += 1

    # 股息
    for cur, d in sorted(tax_summary['dividends']['by_currency'].items()):
        vals = ['股息', cur, d['gross'], d['gross_rmb'], d['tax_withheld'],
                d['tax_withheld_rmb'], d['should_tax_rmb'], d['owed_rmb']]
        for c, v in enumerate(vals, 1):
            apply_cell_style(ws3.cell(row=row, column=c, value=v),
                             NORMAL_FONT, None, CENTER_ALIGN, THIN_BORDER)
        row += 1

    # 利息
    for cur, d in sorted(tax_summary['interests']['by_currency'].items()):
        vals = ['利息/票息', cur, d['total'], d['total_rmb'], 0, 0, d['tax_rmb'], d['tax_rmb']]
        for c, v in enumerate(vals, 1):
            apply_cell_style(ws3.cell(row=row, column=c, value=v),
                             NORMAL_FONT, None, CENTER_ALIGN, THIN_BORDER)
        row += 1

    div_total = tax_summary['dividends']['total_tax_rmb'] + tax_summary['interests']['total_tax_rmb']
    apply_cell_style(ws3.cell(row=row, column=1, value="股息利息税小计"), BOLD_FONT, YELLOW_FILL, CENTER_ALIGN, THIN_BORDER)
    for c in range(2, 7):
        apply_cell_style(ws3.cell(row=row, column=c, value=""), BOLD_FONT, YELLOW_FILL, CENTER_ALIGN, THIN_BORDER)
    apply_cell_style(ws3.cell(row=row, column=8, value=round(div_total, 2)), RED_BOLD, YELLOW_FILL, CENTER_ALIGN, THIN_BORDER)
    row += 2

    # ── 三、合计 ──
    ws3.merge_cells(f'A{row}:H{row}')
    grand = tax_summary['grand_total_rmb']
    ws3.cell(row=row, column=1,
             value=f"三、合计应补缴个人所得税: ¥{grand:,.2f}").font = Font(name='微软雅黑', size=14, bold=True, color='FF0000')
    ws3.row_dimensions[row].height = 35
    row += 2

    # ── 注意事项 ──
    ws3.cell(row=row, column=1, value="⚠️ 重要提醒").font = Font(name='微软雅黑', size=11, bold=True, color='FF0000')
    row += 1
    reminders = [
        "• 申报时间: 次年3月1日至6月30日，逾期按日加收万分之五滞纳金（年化≈18.25%）",
        "• 浮盈不征税: 仅对实际卖出实现的盈利计税",
        "• 跨年度亏损不能结转抵扣",
        "• 美股股息预扣10%(W8BEN)可申请税收抵免，仅需补缴差额",
        "• RSU归属时已按工资薪金计税，卖出时以归属日市价为成本",
        f"• 汇率按{year}年12月31日人民币中间价折算",
        "• 建议保留完整交易记录和完税证明至少5年",
    ]
    for r in reminders:
        ws3.cell(row=row, column=1, value=r).font = NORMAL_FONT
        row += 1

    # 列宽
    for i, w in enumerate([30, 8, 14, 10, 14, 14, 14, 14], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 4: 股息明细 ──
    ws4 = wb.create_sheet("💵 股息利息明细")
    div_d_headers = ['日期', '账户', '类型', '方向', '币种', '金额', '备注']
    for c, h in enumerate(div_d_headers, 1):
        apply_cell_style(ws4.cell(row=1, column=c, value=h),
                         HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)

    r = 2
    for div in fund_flows.get('dividends', []):
        for c, k in enumerate(['date', 'account_name', 'type', 'direction', 'currency', 'amount', 'remark'], 1):
            apply_cell_style(ws4.cell(row=r, column=c, value=div[k]),
                             NORMAL_FONT, None, LEFT_ALIGN, THIN_BORDER)
        r += 1

    for coupon in fund_flows.get('note_coupons', []):
        for c, k in enumerate(['date', 'account_name', 'type', 'direction', 'currency', 'amount', 'remark'], 1):
            apply_cell_style(ws4.cell(row=r, column=c, value=coupon[k]),
                             NORMAL_FONT, None, LEFT_ALIGN, THIN_BORDER)
        r += 1

    # ── Sheet 5: 警告/注意事项 ──
    if warnings:
        ws5 = wb.create_sheet("⚠️ 需确认事项")
        warn_headers = ['代码', '市场', '类型', '日期', '数量', '说明']
        for c, h in enumerate(warn_headers, 1):
            apply_cell_style(ws5.cell(row=1, column=c, value=h),
                             HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)
        for i, w in enumerate(warnings):
            vals = [w['code'], w['market'], w['type'], w['date'], w['qty'], w['message']]
            for c, v in enumerate(vals, 1):
                apply_cell_style(ws5.cell(row=i+2, column=c, value=v),
                                 NORMAL_FONT, YELLOW_FILL, LEFT_ALIGN, THIN_BORDER)

    # 保存
    wb.save(output_path)
    return output_path


# ═══════════════════════════════════════════════════════════════════
# 8. 单年度处理
# ═══════════════════════════════════════════════════════════════════

def process_single_year(input_file, prior_year_file=None, output_file=None,
                        year=None, method='WMA', exchange_rates=None):
    """处理单个年度的账单"""

    if year is None:
        match = re.search(r'(\d{4})', os.path.basename(input_file))
        year = int(match.group(1)) if match else datetime.now().year

    print(f"\n{'='*60}")
    print(f"  📊 处理 {year}年度: {os.path.basename(input_file)}")
    print(f"     方法: {'移动加权平均(WMA)' if method == 'WMA' else '先进先出(FIFO)'}")
    print(f"{'='*60}")

    # 读取
    wb = openpyxl.load_workbook(input_file, data_only=True)

    trading = parse_trading_records(wb['证券-交易流水'])
    assets = parse_asset_movements(wb['证券-资产进出'])
    flows = parse_fund_flows(wb['证券-资金进出'])

    print(f"  交易记录: {len(trading)} 条")
    print(f"  股息相关: {len(flows['dividends'])} 条")
    print(f"  利息相关: {len(flows['interests'])} 条")

    # 期初持仓
    opening_positions = {}

    if prior_year_file:
        print(f"  读取上年文件: {os.path.basename(prior_year_file)}")
        wb_prior = openpyxl.load_workbook(prior_year_file, data_only=True)

        # 处理上年交易以获取准确成本
        prior_trading = parse_trading_records(wb_prior['证券-交易流水'])
        prior_assets = parse_asset_movements(wb_prior['证券-资产进出'])
        prior_flows = parse_fund_flows(wb_prior['证券-资金进出'])
        prior_opening = parse_opening_positions(wb_prior['证券-持仓总览'], '期初')

        _, prior_trackers, _ = calculate_all(
            prior_trading, prior_assets, prior_flows, prior_opening, method
        )

        for key, t in prior_trackers.items():
            if t.P > 0.001:
                code, market, currency = key
                opening_positions[(code, market, currency)] = {
                    'qty': t.P, 'price': t.Q,
                    'market': market, 'currency': currency, 'code': code,
                }
        print(f"  期初持仓(已结转成本): {len(opening_positions)} 个")
    else:
        if '证券-持仓总览' in wb.sheetnames:
            opening_positions = parse_opening_positions(wb['证券-持仓总览'], '期初')
            if opening_positions:
                print(f"  ⚠️ 未提供上年文件，期初成本用市价近似({len(opening_positions)}个)")
            else:
                print(f"  无期初持仓")

    # 汇率
    if exchange_rates is None:
        print(f"  获取汇率...")
        exchange_rates = fetch_exchange_rates(year)
        if not exchange_rates:
            print(f"  ⚠️ 无法自动获取汇率，请手动输入")
            exchange_rates = {'USD': 7.0, 'HKD': 0.9}
            print(f"     使用默认: USD/CNY=7.0, HKD/CNY=0.9")
    else:
        print(f"  使用指定汇率")

    for cur, rate in exchange_rates.items():
        print(f"    {cur}/CNY = {rate:.4f}")

    # 计算
    print(f"  计算 P/Q/R...")
    enriched, trackers, warnings = calculate_all(
        trading, assets, flows, opening_positions, method
    )

    total_R = sum(r['R'] for r in enriched)
    profit_cnt = sum(1 for r in enriched if r['R'] > 0.005)
    loss_cnt = sum(1 for r in enriched if r['R'] < -0.005)
    print(f"  R列合计: {total_R:,.2f} (盈利{profit_cnt}笔/亏损{loss_cnt}笔)")
    print(f"  跟踪证券: {len(trackers)} 个")

    # 税务
    tax = calculate_tax(trackers, flows, exchange_rates, year)

    # 输出
    if output_file is None:
        base = os.path.splitext(os.path.basename(input_file))[0]
        output_file = f"{base}_税务审计底稿_{method}.xlsx"

    write_output(enriched, trackers, tax, flows, warnings,
                 exchange_rates, output_file, year, method)

    print(f"  ✅ 输出: {output_file}")

    # 打印摘要
    print(f"\n  {'─'*40}")
    print(f"  📋 {year}年度 税务摘要")
    print(f"  {'─'*40}")
    for market, mkt in sorted(tax['capital_gains']['by_market'].items()):
        for cur, d in sorted(mkt['currencies'].items()):
            print(f"  {mkt['label']:12s} {cur:4s}  盈亏: {d['pnl']:>12,.2f}  税额: ¥{d['tax_rmb']:>10,.2f}")
    print(f"  {'资本利得税小计':>22s}: ¥{tax['capital_gains']['total_tax_rmb']:>10,.2f}")
    print(f"  {'股息红利税小计':>22s}: ¥{tax['dividends']['total_tax_rmb']:>10,.2f}")
    print(f"  {'利息收入税小计':>22s}: ¥{tax['interests']['total_tax_rmb']:>10,.2f}")
    print(f"  {'─'*40}")
    print(f"  {'合计应补税':>22s}: ¥{tax['grand_total_rmb']:>10,.2f}")
    print(f"  {'='*60}\n")

    if warnings:
        print(f"  ⚠️ {len(warnings)} 项需要确认:")
        for w in warnings[:5]:
            print(f"    - {w['message']}")
        if len(warnings) > 5:
            print(f"    ...还有{len(warnings)-5}项")

    return {
        'year': year,
        'output_file': output_file,
        'tax': tax,
        'total_R': total_R,
        'warnings': warnings,
        'enriched': enriched,
        'trackers': trackers,
    }


# ═══════════════════════════════════════════════════════════════════
# 9. 批量处理所有年度
# ═══════════════════════════════════════════════════════════════════

def process_all_years(directory, method='WMA', exchange_rates=None):
    """批量处理目录下所有年度账单"""
    import glob

    pattern = os.path.join(directory, '*年度账单*.xlsx')
    files = sorted(glob.glob(pattern))

    if not files:
        print(f"未找到年度账单文件: {pattern}")
        return None

    years = []
    for f in files:
        match = re.search(r'(\d{4})', os.path.basename(f))
        if match:
            years.append((int(match.group(1)), f))

    years.sort()

    if exchange_rates is None and years:
        # 尝试获取最新年度的汇率
        exchange_rates = fetch_exchange_rates(years[-1][0])

    results = []
    for i, (year, filepath) in enumerate(years):
        prior = years[i-1][1] if i > 0 else None
        output = os.path.join(directory, 'output',
                              f"{year}_税务审计底稿_{method}.xlsx")
        os.makedirs(os.path.dirname(output), exist_ok=True)

        result = process_single_year(
            filepath, prior, output, year, method, exchange_rates
        )
        results.append(result)

    # 生成汇总报告
    if results:
        _write_summary_report(results, directory, method, exchange_rates)

    return results


def _write_summary_report(results, directory, method, exchange_rates):
    """生成多年度汇总报告"""
    output = os.path.join(directory, 'output', f'所有年度汇总_{method}.xlsx')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "年度汇总"

    # Title
    ws.merge_cells('A1:I1')
    ws.cell(row=1, column=1, value="境外所得税 — 多年度汇总报告").font = TITLE_FONT
    ws.row_dimensions[1].height = 35

    row = 3
    headers = ['年度', '资本利得税(RMB)', '股息红利税(RMB)', '利息税(RMB)',
               '合计应补税(RMB)', '港美股盈亏', '美股盈亏', '总交易笔数', '备注']
    for c, h in enumerate(headers, 1):
        apply_cell_style(ws.cell(row=row, column=c, value=h),
                         HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)
    row += 1

    grand_total = 0
    for r in results:
        t = r['tax']
        # 统计港美股盈亏
        hk_pnl = sum(d['pnl'] for mkt_name, mkt in t['capital_gains']['by_market'].items()
                    if '港股' in mkt['label'] for d in mkt['currencies'].values())
        us_pnl = sum(d['pnl'] for mkt_name, mkt in t['capital_gains']['by_market'].items()
                    if '美股' in mkt['label'] for d in mkt['currencies'].values())

        vals = [
            r['year'],
            t['capital_gains']['total_tax_rmb'],
            t['dividends']['total_tax_rmb'],
            t['interests']['total_tax_rmb'],
            t['grand_total_rmb'],
            round(hk_pnl, 2),
            round(us_pnl, 2),
            len(r['enriched']),
            f"⚠️{len(r['warnings'])}项" if r['warnings'] else "✓"
        ]
        for c, v in enumerate(vals, 1):
            apply_cell_style(ws.cell(row=row, column=c, value=v),
                             NORMAL_FONT, None, CENTER_ALIGN, THIN_BORDER)

        grand_total += t['grand_total_rmb']
        row += 1

    # Total row
    apply_cell_style(ws.cell(row=row, column=1, value="合计"), BOLD_FONT, YELLOW_FILL, CENTER_ALIGN, THIN_BORDER)
    apply_cell_style(ws.cell(row=row, column=5, value=round(grand_total, 2)),
                     RED_BOLD, YELLOW_FILL, CENTER_ALIGN, THIN_BORDER)
    for c in range(2, 9):
        apply_cell_style(ws.cell(row=row, column=c, value=""), BOLD_FONT, YELLOW_FILL, CENTER_ALIGN, THIN_BORDER)

    # Column widths
    widths = [8, 16, 16, 14, 16, 14, 14, 12, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(output)
    print(f"\n📊 多年度汇总报告: {output}")
    return output


# ═══════════════════════════════════════════════════════════════════
# 10. CLI入口
# ═══════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description=f'富途 CRS 境外所得税计算器 v{VERSION}',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
使用示例:
  # 单年度
  python3 futu_tax_calculator.py 2025_年度账单.xlsx 2024_年度账单.xlsx

  # 批量处理
  python3 futu_tax_calculator.py --all-years .

  # 指定方法和汇率
  python3 futu_tax_calculator.py 2025_年度账单.xlsx --method FIFO --usd-rate 7.2 --hkd-rate 0.92
        """
    )
    parser.add_argument('input_file', nargs='?', help='当年度账单 Excel')
    parser.add_argument('prior_year_file', nargs='?', help='上年度账单(可选)')
    parser.add_argument('-o', '--output', help='输出文件路径')
    parser.add_argument('--year', type=int, help='纳税年度')
    parser.add_argument('--method', choices=['WMA', 'FIFO'], default='WMA',
                        help='成本计算方法: WMA=移动加权平均(默认), FIFO=先进先出')
    parser.add_argument('--usd-rate', type=float, help='USD/CNY 汇率')
    parser.add_argument('--hkd-rate', type=float, help='HKD/CNY 汇率')
    parser.add_argument('--all-years', metavar='DIR',
                        help='批量处理目录下所有年度账单')

    args = parser.parse_args()

    # 汇率
    exchange_rates = {}
    if args.usd_rate:
        exchange_rates['USD'] = args.usd_rate
    if args.hkd_rate:
        exchange_rates['HKD'] = args.hkd_rate
    if not exchange_rates:
        exchange_rates = None

    # 批量模式
    if args.all_years:
        process_all_years(args.all_years, args.method, exchange_rates)
        return 0

    # 单年度模式
    if not args.input_file:
        parser.print_help()
        return 1

    process_single_year(
        args.input_file,
        args.prior_year_file,
        args.output,
        args.year,
        args.method,
        exchange_rates,
    )
    return 0


if __name__ == '__main__':
    sys.exit(main())
